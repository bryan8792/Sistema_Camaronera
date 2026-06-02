import json
import os
from datetime import datetime
from django.conf import settings
from io import BytesIO
import xlsxwriter
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import get_template
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.template import RequestContext, loader
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.views.generic.base import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from app_contabilidad_planCuentas.forms import PlanCuentaForm, EncabezadoCuentasPlanCuentaForm
from app_contabilidad_planCuentas.models import PlanCuenta, Folder, EncabezadoCuentasPlanCuenta, \
    DetalleCuentasPlanCuenta
from app_empresa.app_reg_empresa.models import Empresa
from collections import defaultdict
from django.views.generic import ListView
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Sum, F, Q
from weasyprint import HTML, CSS
from django.db.models.functions import Coalesce
from datetime import datetime, timedelta
from decimal import Decimal
import logging
from django.forms.models import model_to_dict
import re

logger = logging.getLogger(__name__)


def _inicializar():
    Folder.objects.all().delete()
    folder_raiz = Folder(
        name='Plan de Cuentas'
    )
    folder_raiz.save()
    folder_raiz2 = Folder(
        name='Plan de Cuentas V2'
    )
    folder_raiz2.save()
    folder = Folder(
        name='Docs',
        parent=folder_raiz
    )
    folder.save()
    folder_trab = Folder(
        name='Trabajo',
        parent=folder
    )
    folder_trab.save()

    folder = Folder(
        name='Música',
        parent=folder_raiz
    )
    folder.save()
    folder_photos = Folder(
        name='Fotos',
        parent=folder_raiz
    )
    folder_photos.save()
    folder = Folder(
        name='Vacaciones',
        parent=folder_photos
    )
    folder.save()
    folder = Folder(
        name='Trabajo',
        parent=folder_photos
    )
    folder.save()


# CREAR PLAN DE CUENTAS
class crearPlanCuentaView(CreateView):
    model = PlanCuenta
    form_class = PlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_BIO')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        parent_id = self.request.GET.get('parent_id')
        if parent_id:
            try:
                padre = PlanCuenta.objects.get(pk=parent_id)
                hijos = PlanCuenta.objects.filter(parentId=padre).order_by('codigo')
                if hijos.exists():
                    ultimo_codigo = hijos.last().codigo
                    nuevo_codigo = str(int(ultimo_codigo) + 1).zfill(len(ultimo_codigo))
                else:
                    nuevo_codigo = padre.codigo + '01'

                initial['codigo'] = nuevo_codigo
                initial['parentId'] = padre.id

            except PlanCuenta.DoesNotExist:
                pass

            siglas = self.request.GET.get('siglas', 'BIO')
            empresa = Empresa.objects.filter(siglas=siglas).first()
            if empresa:
                initial['empresa'] = empresa.id
        return initial

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'create':
                form = self.get_form()
                if form.is_valid():
                    plan = form.save()
                    data = model_to_dict(plan)
                else:
                    data['error'] = form.errors.as_json()
            else:
                data['error'] = 'No ha ingresado a ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Plan de Cuenta Empresa BIO'
        context['action'] = 'create'
        context['list_url'] = self.success_url
        context['codigo_padre'] = self.request.GET.get('codigo_padre', '')
        context['parent_id'] = self.request.GET.get('parent_id', '')
        return context


# ACTUALIZAR PLAN DE CUENTAS
class actualizarPlanCuentaView(UpdateView):
    model = PlanCuenta
    form_class = PlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_BIO')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'edit':
                form = self.get_form()
                if form.is_valid():
                    plan = form.save()
                    data = model_to_dict(plan)
                else:
                    data['error'] = form.errors.as_json()
            else:
                data['error'] = 'No ha ingresado a ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Actualizar Plan de Cuenta'
        context['action'] = 'edit'
        context['list_url'] = self.success_url

        # Mostrar también el código del padre (si aplica)
        if self.object.parentId:
            context['codigo_padre'] = self.object.parentId.codigo
            context['parent_id'] = self.object.parentId.id
        else:
            context['codigo_padre'] = ''
            context['parent_id'] = ''

        return context


# ELIMINAR PLAN DE CUENTAS
class eliminarPlanCuentaView(DeleteView):
    model = PlanCuenta
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_eliminar.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_BIO')
    url_redirect = success_url

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.object.delete()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Eliminar Plan de Cuenta'
        context['list_url'] = self.success_url
        return context


# LISTAR PLAN DE CUENTAS
class listarPlanCuentaBIOView(TemplateView):
    model = PlanCuenta
    template_name = 'app_contabilidad_planCuentas/parts_Plan/foldersBIO.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        # _inicializar()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'searchdataBIO':
                data = []
                empresa = request.POST['empresa']
                queryset = PlanCuenta.objects.all()
                queryset = queryset.filter(empresa__siglas=empresa).order_by('codigo')
                for i in queryset:
                    data.append(i.toJSON())

            # if action == 'get_children':
            #     data = []
            #     empresa_sigla = request.POST.get('empresa')
            #     parent_id = request.POST.get('parent_id')  # puede ser vacío o un ID
            #
            #     # Raíz (padre null)
            #     if parent_id == "#":
            #         cuentas = PlanCuenta.objects.filter(parentId__isnull=True, empresa__siglas=empresa_sigla).order_by(
            #             'codigo')
            #     else:
            #         cuentas = PlanCuenta.objects.filter(parentId_id=parent_id, empresa__siglas=empresa_sigla).order_by(
            #             'codigo')
            #
            #     for cuenta in cuentas:
            #         has_children = PlanCuenta.objects.filter(parentId=cuenta).exists()
            #         data.append({
            #             "id": str(cuenta.id),
            #             "text": f"<strong>{cuenta.codigo}</strong> &nbsp; {cuenta.nombre}",
            #             "children": has_children,  # Esto le dice a JSTree si debe seguir cargando al expandir
            #             "data": {
            #                 "nivel": cuenta.nivel
            #             }
            #         })

            if action == 'get_plan':
                data = []
                empresa_sigla = request.POST.get('empresa')
                parent_id = request.POST.get('parent_id')

                if parent_id == "#":
                    cuentas = PlanCuenta.objects.filter(parentId__isnull=True, empresa__siglas=empresa_sigla).order_by(
                        'codigo')
                else:
                    cuentas = PlanCuenta.objects.filter(parentId_id=parent_id, empresa__siglas=empresa_sigla).order_by(
                        'codigo')

                for cuenta in cuentas:
                    has_children = PlanCuenta.objects.filter(parentId=cuenta).exists()
                    icon_type = 'jstree-folder' if has_children else 'jstree-file'

                    data.append({
                        "id": str(cuenta.id),
                        "text": f"""
                            <span class='context-menu-one'>
                                <strong>{cuenta.codigo}</strong> &nbsp; {cuenta.nombre}
                            </span>
                            <span class='nivel-info' style="position: absolute; right: 0;"><b>Nivel {cuenta.nivel}</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span> 
                        """,
                        "children": has_children,
                        "icon": icon_type
                    })


            elif action == 'upload_excel':
                print('llego a Upload excell empezo a recorrer en el python desde ajax')
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]

                    for row in range(2, excel.max_row + 1):
                        name_empresa = excel.cell(row=row, column=7).value

                        # Validar si la empresa existe por sus siglas
                        empresa, created = Empresa.objects.get_or_create(
                            siglas=name_empresa,
                            defaults={
                                'nombre': name_empresa  # Ajustar según sea necesario si las siglas son distintas
                            }
                        )

                        print(
                            f"{'Nueva empresa creada' if created else 'Empresa existente'}: {empresa.nombre} con siglas {empresa.siglas}")

                        # Buscar o inicializar la cuenta contable
                        codigo_cuenta = excel.cell(row=row, column=2).value
                        product, created = PlanCuenta.objects.get_or_create(
                            empresa=empresa,
                            codigo=codigo_cuenta,
                            defaults={
                                'nombre': excel.cell(row=row, column=3).value,
                                'nivel': int(excel.cell(row=row, column=4).value),
                                'tipo_cuenta': excel.cell(row=row, column=5).value,
                                'periodo': int(excel.cell(row=row, column=8).value)
                            }
                        )

                        # Si la cuenta ya existe, actualizarla con los nuevos valores
                        if not created:
                            product.nombre = excel.cell(row=row, column=3).value
                            product.nivel = int(excel.cell(row=row, column=4).value)
                            product.tipo_cuenta = excel.cell(row=row, column=5).value
                            product.periodo = int(excel.cell(row=row, column=8).value)

                        # Manejo del padre (cuenta superior)
                        codigo_padre = excel.cell(row=row, column=6).value
                        if codigo_padre:
                            try:
                                padre = PlanCuenta.objects.get(codigo=codigo_padre, empresa=empresa)
                                product.parentId = padre
                            except PlanCuenta.DoesNotExist:
                                # Si no existe el padre, mostrar un mensaje o manejar el error
                                print(
                                    f"Cuenta padre con código {codigo_padre} no existe para la empresa {empresa.siglas}")

                        product.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Listado de Plan de Cuentas'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        folders = Folder.objects.filter(parent=None)
        context['folders'] = folders
        return context


# PLAN DE CUENTAS DE LA EMPRESA PSM CREAR
class crearPlanCuentaPSMView(CreateView):
    model = PlanCuenta
    form_class = PlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_crear_psm.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_PSM')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Plan de Cuenta'
        context['action'] = 'add'
        context['list_url'] = self.success_url
        return context


# ACTUALIZAR PLAN DE CUENTAS
class actualizarPlanCuentaPSMView(UpdateView):
    model = PlanCuenta
    form_class = PlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_crear_psm.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_PSM')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Actualizar Plan de Cuenta'
        context['action'] = 'edit'
        context['list_url'] = self.success_url
        return context


# ELIMINAR PLAN DE CUENTAS
class eliminarPlanCuentaPSMView(DeleteView):
    model = PlanCuenta
    form_class = PlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/parts_Plan/planCuenta_crear_psm.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_PSM')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Eliminar Plan de Cuenta'
        context['list_url'] = self.success_url
        return context


#     LISTAR PLAN DE CUENTAS PSM
# class listarPlanCuentaPSMView(TemplateView):
#     model = PlanCuenta
#     template_name = 'app_contabilidad_planCuentas/parts_Plan/foldersPSM.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         # _inicializar()
#         return super().dispatch(request, *args, **kwargs)
#
#     # def post(self, request, *args, **kwargs):
#     #     data = {}
#     #     try:
#     #         data = PlanCuenta.objects.get(pk=request.POST['id']).toJSON()
#     #         print('data')
#     #         print(data)
#     #     except Exception as e:
#     #         data['error'] = str(e)
#     #     return JsonResponse(data)
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#         action = request.POST['action']
#         try:
#             if action == 'searchdataPSM':
#                 data = []
#                 empresa = request.POST['empresa']
#                 queryset = PlanCuenta.objects.all()
#                 queryset = queryset.filter(empresa__siglas=empresa).order_by('codigo')
#                 for i in queryset:
#                     data.append(i.toJSON())
#
#             elif action == 'upload_excel':
#                 print('llego a Upload excell empezo a recorrer en el python desde ajax')
#                 with transaction.atomic():
#                     archive = request.FILES['archive']
#                     workbook = load_workbook(filename=archive, data_only=True)
#                     excel = workbook[workbook.sheetnames[0]]
#
#                     for row in range(2, excel.max_row + 1):
#                         name_empresa = excel.cell(row=row, column=7).value
#
#                         # Validar si la empresa existe por sus siglas
#                         empresa, created = Empresa.objects.get_or_create(
#                             siglas=name_empresa,
#                             defaults={
#                                 'nombre': name_empresa  # Ajustar según sea necesario si las siglas son distintas
#                             }
#                         )
#
#                         print(
#                             f"{'Nueva empresa creada' if created else 'Empresa existente'}: {empresa.nombre} con siglas {empresa.siglas}")
#
#                         # Buscar o inicializar la cuenta contable
#                         codigo_cuenta = excel.cell(row=row, column=2).value
#                         product, created = PlanCuenta.objects.get_or_create(
#                             empresa=empresa,
#                             codigo=codigo_cuenta,
#                             defaults={
#                                 'nombre': excel.cell(row=row, column=3).value,
#                                 'nivel': int(excel.cell(row=row, column=4).value),
#                                 'tipo_cuenta': excel.cell(row=row, column=5).value,
#                                 'periodo': int(excel.cell(row=row, column=8).value)
#                             }
#                         )
#
#                         # Si la cuenta ya existe, actualizarla con los nuevos valores
#                         if not created:
#                             product.nombre = excel.cell(row=row, column=3).value
#                             product.nivel = int(excel.cell(row=row, column=4).value)
#                             product.tipo_cuenta = excel.cell(row=row, column=5).value
#                             product.periodo = int(excel.cell(row=row, column=8).value)
#
#                         # Manejo del padre (cuenta superior)
#                         codigo_padre = excel.cell(row=row, column=6).value
#                         if codigo_padre:
#                             try:
#                                 padre = PlanCuenta.objects.get(codigo=codigo_padre, empresa=empresa)
#                                 product.parentId = padre
#                             except PlanCuenta.DoesNotExist:
#                                 # Si no existe el padre, mostrar un mensaje o manejar el error
#                                 print(
#                                     f"Cuenta padre con código {codigo_padre} no existe para la empresa {empresa.siglas}")
#
#                         product.save()
#             else:
#                 data['error'] = 'Ha ocurrido un error'
#         except Exception as e:
#             data['error'] = str(e)
#         return HttpResponse(json.dumps(data), content_type='application/json')
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Listado de Plan de Cuentas'
#         planCuenta = PlanCuenta.objects.filter(parentId=None)
#         context['planCuenta'] = planCuenta
#         folders = Folder.objects.filter(parent=None)
#         context['folders'] = folders
#         return context

class listarPlanCuentaPSMView(TemplateView):
    model = PlanCuenta
    template_name = 'app_contabilidad_planCuentas/parts_Plan/foldersPSM.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        # _inicializar()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'searchdataPSM':
                data = []
                empresa = request.POST['empresa']
                queryset = PlanCuenta.objects.all()
                queryset = queryset.filter(empresa__siglas=empresa).order_by('codigo')
                for i in queryset:
                    data.append(i.toJSON())

            # if action == 'get_children':
            #     data = []
            #     empresa_sigla = request.POST.get('empresa')
            #     parent_id = request.POST.get('parent_id')  # puede ser vacío o un ID
            #
            #     # Raíz (padre null)
            #     if parent_id == "#":
            #         cuentas = PlanCuenta.objects.filter(parentId__isnull=True, empresa__siglas=empresa_sigla).order_by(
            #             'codigo')
            #     else:
            #         cuentas = PlanCuenta.objects.filter(parentId_id=parent_id, empresa__siglas=empresa_sigla).order_by(
            #             'codigo')
            #
            #     for cuenta in cuentas:
            #         has_children = PlanCuenta.objects.filter(parentId=cuenta).exists()
            #         data.append({
            #             "id": str(cuenta.id),
            #             "text": f"<strong>{cuenta.codigo}</strong> &nbsp; {cuenta.nombre}",
            #             "children": has_children,  # Esto le dice a JSTree si debe seguir cargando al expandir
            #             "data": {
            #                 "nivel": cuenta.nivel
            #             }
            #         })

            if action == 'get_plan':
                data = []
                empresa_sigla = request.POST.get('empresa')
                parent_id = request.POST.get('parent_id')

                if parent_id == "#":
                    cuentas = PlanCuenta.objects.filter(parentId__isnull=True, empresa__siglas=empresa_sigla).order_by(
                        'codigo')
                else:
                    cuentas = PlanCuenta.objects.filter(parentId_id=parent_id, empresa__siglas=empresa_sigla).order_by(
                        'codigo')

                for cuenta in cuentas:
                    has_children = PlanCuenta.objects.filter(parentId=cuenta).exists()
                    icon_type = 'jstree-folder' if has_children else 'jstree-file'

                    data.append({
                        "id": str(cuenta.id),
                        "text": f"""
                            <span class='context-menu-one'>
                                <strong>{cuenta.codigo}</strong> &nbsp; {cuenta.nombre}
                            </span>
                            <span class='nivel-info' style="position: absolute; right: 0;"><b>Nivel {cuenta.nivel}</b>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span> 
                        """,
                        "children": has_children,
                        "icon": icon_type
                    })


            elif action == 'upload_excel':
                print('llego a Upload excell empezo a recorrer en el python desde ajax')
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]

                    for row in range(2, excel.max_row + 1):
                        name_empresa = excel.cell(row=row, column=7).value

                        # Validar si la empresa existe por sus siglas
                        empresa, created = Empresa.objects.get_or_create(
                            siglas=name_empresa,
                            defaults={
                                'nombre': name_empresa  # Ajustar según sea necesario si las siglas son distintas
                            }
                        )

                        print(
                            f"{'Nueva empresa creada' if created else 'Empresa existente'}: {empresa.nombre} con siglas {empresa.siglas}")

                        # Buscar o inicializar la cuenta contable
                        codigo_cuenta = excel.cell(row=row, column=2).value
                        product, created = PlanCuenta.objects.get_or_create(
                            empresa=empresa,
                            codigo=codigo_cuenta,
                            defaults={
                                'nombre': excel.cell(row=row, column=3).value,
                                'nivel': int(excel.cell(row=row, column=4).value),
                                'tipo_cuenta': excel.cell(row=row, column=5).value,
                                'periodo': int(excel.cell(row=row, column=8).value)
                            }
                        )

                        # Si la cuenta ya existe, actualizarla con los nuevos valores
                        if not created:
                            product.nombre = excel.cell(row=row, column=3).value
                            product.nivel = int(excel.cell(row=row, column=4).value)
                            product.tipo_cuenta = excel.cell(row=row, column=5).value
                            product.periodo = int(excel.cell(row=row, column=8).value)

                        # Manejo del padre (cuenta superior)
                        codigo_padre = excel.cell(row=row, column=6).value
                        if codigo_padre:
                            try:
                                padre = PlanCuenta.objects.get(codigo=codigo_padre, empresa=empresa)
                                product.parentId = padre
                            except PlanCuenta.DoesNotExist:
                                # Si no existe el padre, mostrar un mensaje o manejar el error
                                print(
                                    f"Cuenta padre con código {codigo_padre} no existe para la empresa {empresa.siglas}")

                        product.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Listado de Plan de Cuentas'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        folders = Folder.objects.filter(parent=None)
        context['folders'] = folders
        return context



# EXPORTAR A EXCELL PLAN DE CUENTAS
class PlanExportExcelBIOView(View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                'id': 10,
                'CODIGO': 15,
                'NOMBRE': 75,
                'NIVEL': 20,
                'GEN_DET': 20,
                'CTA_MAY': 20,
                'EMPRESA': 20,
                'PERIODO': 20,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('plan_cuentas')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            row_format2 = workbook.add_format({'align': 'left', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=0, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            queryset = PlanCuenta.objects.all()
            queryset = queryset.filter(empresa__siglas='BIO').order_by('id')
            for product in queryset:
                worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 1, product.codigo, row_format2)
                worksheet.write(row, 2, product.nombre, row_format2)
                worksheet.write(row, 3, product.nivel, row_format)
                worksheet.write(row, 4, product.tipo_cuenta, row_format)
                worksheet.write(row, 5, product.code_parent(), row_format2)
                worksheet.write(row, 6, product.empresa.siglas, row_format)
                worksheet.write(row, 7, product.periodo, row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Plan_Cuentas_BIO_{}.xlsx"'.format(
                datetime.now().date().strftime('%d_%m_%Y'))
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('app_planCuentas:listar_planCuenta'))


class PlanExportExcelPSMView(View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                'id': 10,
                'CODIGO': 15,
                'NOMBRE': 75,
                'NIVEL': 20,
                'GEN_DET': 20,
                'CTA_MAY': 20,
                'EMPRESA': 20,
                'PERIODO': 20,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('plan_cuentas')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            row_format2 = workbook.add_format({'align': 'left', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=0, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            queryset = PlanCuenta.objects.all()
            queryset = queryset.filter(empresa__siglas='PSM').order_by('id')
            for product in queryset:
                worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 1, product.codigo, row_format2)
                worksheet.write(row, 2, product.nombre, row_format2)
                worksheet.write(row, 3, product.nivel, row_format)
                worksheet.write(row, 4, product.tipo_cuenta, row_format)
                worksheet.write(row, 5, product.code_parent(), row_format2)
                worksheet.write(row, 6, product.empresa.siglas, row_format)
                worksheet.write(row, 7, product.periodo, row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Plan_Cuentas_PSM_{}.xlsx"'.format(
                datetime.now().date().strftime('%d_%m_%Y'))
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('app_planCuentas:listar_planCuenta'))


# LISTAR TRANSACCIONES DEL PLAN DE CUENTAS
class listarTransaccionPlanView(ListView):
    model = EncabezadoCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_listar.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in EncabezadoCuentasPlanCuenta.objects.filter(reg_control__exact='RT',
                                                                    empresa__siglas__exact='PSM'):
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Transaccion de Plan de Cuentas Empresa PSM'
        context['title'] = 'Transaccion de Plan de Cuentas Empresa PSM'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


class listarTransaccionPlanBIOView(ListView):
    model = EncabezadoCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_listar_bio.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata_bio':
                data = []
                for i in EncabezadoCuentasPlanCuenta.objects.filter(reg_control__exact='RT', empresa__siglas__exact='BIO'):
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Transacciones Sin ATS del Plan de Cuentas Empresa BIO'
        context['title'] = 'Transaccion de Plan de Cuentas Empresa BIO'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


# CREAR TRANSACCIONES DEL PLAN DE CUENTAS
class crearTransaccionPlanView(CreateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_plan':
                data = []
                empresa = request.POST['empresa']
                print('empresa de search plan')
                print(empresa)
                queryset = PlanCuenta.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                queryset = queryset.filter(empresa__siglas=empresa).exclude(id__in=ids_exclude)
                # if len(ids_exclude):
                #     queryset = queryset.filter().exclude(id__in=ids_exclude)
                for i in queryset:
                    item = i.toJSON()
                    item['detalle'] = ""
                    data.append(item)

            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST['ids'])
                term = request.POST['term'].strip()
                data.append({'codigo': term, 'text': term})
                plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
                for i in plan_detail[0:50]:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = i.nombre
                    data.append(item)

            elif action == 'create':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = EncabezadoCuentasPlanCuenta()
                    encabezado.codigo = request.POST['codigo']
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.empresa_id = request.POST['empresa']
                    encabezado.save()
                    for i in items:
                        cuerpo = DetalleCuentasPlanCuenta()
                        cuerpo.encabezadocuentaplan_id = encabezado.pk
                        cuerpo.cuenta_id = int(i['id'])
                        cuerpo.detalle = i['detalle']
                        cuerpo.debe = int(i['debe']) if i.get('debe') else 0
                        cuerpo.haber = int(i['haber']) if i.get('haber') else 0
                        cuerpo.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'create'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.all()
        context['planCuenta2'] = planCuenta2
        context['det'] = []
        return context


# class TransaccionInvoicePdfView(View):
#
#     def get(self, request, *args, **kwargs):
#         try:
#             template = get_template('app_reportes/reporte_transaccion.html')
#
#             encabezado = EncabezadoCuentasPlanCuenta.objects.get(pk=self.kwargs['pk'])
#             detalle = DetalleCuentasPlanCuenta.objects.filter(encabezadocuentaplan=encabezado)
#
#             print('encabezado')
#             print(encabezado)
#             print("detalle")
#             print(detalle)
#
#             # Acceso correcto a empresa según tu modelo
#             empresa = getattr(encabezado, 'empresa', None)
#             empresa_nombre = empresa.nombre if empresa else ''  # <-- CORREGIDO
#
#             # Totales calculados en Python
#             total_debe = detalle.aggregate(total=Sum('debe'))['total'] or 0
#             total_haber = detalle.aggregate(total=Sum('haber'))['total'] or 0
#
#             context = {
#                 'sale': encabezado,
#                 'comp': {
#                     'name': 'INDUSTRIA PESQUERA',
#                     'address': 'MACHALA - EL ORO - ECUADOR',
#                     'numero': '(072) 920 371',
#                     'comprobante': 'COMPROBANTE DE INGRESOS DE PRODUCTOS'
#                 },
#                 'icon': '{}{}'.format(settings.MEDIA_URL, 'logo.png'),
#                 'empresa': empresa_nombre,
#                 'detalle_fac': detalle,
#                 'total_debe': total_debe,
#                 'total_haber': total_haber,
#             }
#
#             html = template.render(context)
#             css_url = os.path.join(settings.BASE_DIR, 'static/lib/bootstrap-4.4.1-dist/css/bootstrap.min.css')
#             pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
#
#             return HttpResponse(pdf, content_type='application/pdf')
#         except Exception as e:
#             print("Error en PDF:", e)
#             return HttpResponseRedirect(
#                 reverse_lazy('app_planCuentas:reporte_pdf', kwargs={'pk': self.kwargs['pk']})
#             )


class TransaccionInvoicePdfView(View):

    def get(self, request, *args, **kwargs):
        try:
            encabezado = get_object_or_404(
                EncabezadoCuentasPlanCuenta.objects.select_related("empresa"),
                pk=self.kwargs['pk']
            )

            detalles = (
                DetalleCuentasPlanCuenta.objects
                .filter(encabezadocuentaplan=encabezado)
                .select_related("cuenta")  # ⚡ reduce queries
                .order_by("id")
            )

            totales = detalles.aggregate(
                total_debe=Sum("debe"),
                total_haber=Sum("haber")
            )
            total_debe = totales["total_debe"] or 0
            total_haber = totales["total_haber"] or 0
            diferencia = total_debe - total_haber

            context = {
                "encabezado": encabezado,
                "detalles": detalles,
                "empresa": encabezado.empresa,
                "total_debe": total_debe,
                "total_haber": total_haber,
                "diferencia": diferencia,
                "icon": f"{settings.MEDIA_URL}logo.png",
            }

            template = get_template("app_reportes/reporte_transaccion.html")
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, "static/lib/bootstrap-4.4.1-dist/css/bootstrap.min.css")
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type="application/pdf")

        except Exception as e:
            print(" Error en PDF:", e)
            return HttpResponseRedirect(reverse_lazy("app_planCuentas:reporte_pdf", kwargs={"pk": self.kwargs["pk"]}))



# # CREAR TRANSACCIONES DEL PLAN DE CUENTAS
# class crearTransaccionPlanBIOView(CreateView):
#     model = EncabezadoCuentasPlanCuenta
#     form_class = EncabezadoCuentasPlanCuentaForm
#     template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crearBIO.html'
#     success_url = reverse_lazy('app_planCuentas:listar_transaccionPlan_bio')
#     url_redirect = success_url
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             action = request.POST['action']
#             if action == 'search_plan':
#                 data = []
#                 empresa = request.POST['empresa']
#                 print('empresa de search plan')
#                 print(empresa)
#                 queryset = PlanCuenta.objects.all()
#                 ids_exclude = json.loads(request.POST['ids'])
#                 queryset = queryset.filter(empresa__siglas=empresa).exclude(id__in=ids_exclude)
#                 # if len(ids_exclude):
#                 #     queryset = queryset.filter().exclude(id__in=ids_exclude)
#                 for i in queryset:
#                     item = i.toJSON()
#                     item['detalle'] = ""
#                     data.append(item)
#
#             elif action == 'search_autocomplete':
#                 data = []
#                 ids_exclude = json.loads(request.POST['ids'])
#                 term = request.POST['term'].strip()
#                 data.append({'codigo': term, 'text': term})
#                 plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
#                 for i in plan_detail[0:50]:
#                     item = i.toJSON()
#                     item['codigo'] = i.codigo
#                     item['text'] = i.nombre
#                     data.append(item)
#
#
#             elif action == 'obtener_ultima_secuencia':
#                 mes = request.POST.get('mes')
#                 tipo = request.POST.get('tipo')
#                 patron_codigo = f"{mes}{tipo}%"
#                 ultimo_encabezado = EncabezadoCuentasPlanCuenta.objects.filter(
#                     codigo__ilike=patron_codigo
#                 ).order_by('-codigo').first()
#                 if ultimo_encabezado:
#                     ultima_secuencia = int(ultimo_encabezado.codigo[-3:])
#                     data['secuencia'] = ultima_secuencia
#                 else:
#                     data['secuencia'] = 0
#
#             elif action == 'create':
#                 with transaction.atomic():
#                     items = json.loads(request.POST['items'])
#                     encabezado = EncabezadoCuentasPlanCuenta()
#                     encabezado.codigo = request.POST['codigo']
#                     encabezado.tip_cuenta = request.POST['tip_cuenta']
#                     encabezado.fecha = request.POST['fecha']
#                     encabezado.comprobante = request.POST['comprobante']
#                     encabezado.descripcion = request.POST['descripcion']
#                     encabezado.direccion = request.POST['direccion']
#                     encabezado.empresa_id = request.POST['empresa']
#                     encabezado.save()
#                     for i in items:
#                         cuerpo = DetalleCuentasPlanCuenta()
#                         cuerpo.encabezadocuentaplan_id = encabezado.pk
#                         cuerpo.cuenta_id = int(i['id'])
#                         cuerpo.detalle = i['detalle']
#                         cuerpo.debe = int(i['debe']) if i.get('debe') else 0
#                         cuerpo.haber = int(i['haber']) if i.get('haber') else 0
#                         cuerpo.save()
#             else:
#                 data['error'] = 'Ha ocurrido un error'
#         except Exception as e:
#             data['error'] = 'el error es : ' + str(e)
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Formulario de Registro de Transacción'
#         context['list_url'] = self.success_url
#         context['action'] = 'create'
#         planCuenta = PlanCuenta.objects.filter(parentId=None)
#         context['planCuenta'] = planCuenta
#         planCuenta2 = PlanCuenta.objects.all()
#         context['planCuenta2'] = planCuenta2
#         context['empresa'] = 'BIO'
#         context['det'] = []
#
#         try:
#             empresa_bio = Empresa.objects.get(siglas='BIO')
#             # Modificar el formulario para preseleccionar la empresa BIO
#             form = self.get_form()
#             form.fields['empresa'].initial = empresa_bio.id
#             context['form'] = form
#         except Exception as e:
#             print(f"Error al preseleccionar empresa BIO: {e}")
#
#         return context


# EDITAR TRANSACCIONES DEL PLAN DE CUENTAS

# CREAR TRANSACCIONES DEL PLAN DE CUENTAS

# CREAR TRANSACCIONES DEL PLAN DE CUENTAS - VERSIÓN CORREGIDA PARA EL ERROR DE TIPO
class crearTransaccionPlanBIOView(CreateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crearBIO.html'
    success_url = reverse_lazy('app_planCuentas:listar_transaccionPlan_bio')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']

            if action == 'search_plan':
                return self.search_plan_improved(request)

            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST.get('ids', '[]'))
                term = request.POST.get('term', '').strip()

                # Agregar el término de búsqueda como primera opción
                data.append({'codigo': term, 'text': term, 'id': None})

                # Buscar cuentas que coincidan con el término
                plan_detail = PlanCuenta.objects.filter(
                    Q(nombre__icontains=term) | Q(codigo__icontains=term),
                    empresa__siglas__exact='BIO'
                ).exclude(id__in=ids_exclude).order_by('codigo')[:50]

                for i in plan_detail:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = f"{i.codigo} - {i.nombre}"
                    item['id'] = int(i.id) if i.id else None
                    data.append(item)

            elif action == 'obtener_ultima_secuencia':
                mes = request.POST.get('mes')
                tipo = request.POST.get('tipo')

                print(f"Buscando secuencia para mes={mes}, tipo={tipo}")

                try:
                    patron_mes = mes.lstrip('0')
                    patron1 = f"{mes}{tipo}"
                    patron2 = f"{patron_mes}{tipo}"

                    encabezados = EncabezadoCuentasPlanCuenta.objects.filter(
                        Q(codigo__startswith=patron1) | Q(codigo__startswith=patron2)
                    ).order_by('-codigo')

                    ultima_secuencia = 0
                    if encabezados.exists():
                        for encabezado in encabezados:
                            codigo = str(encabezado.codigo) if encabezado.codigo is not None else ""
                            print(f"Analizando código: {codigo}")

                            match = re.search(r'(\d{1,2})(\d)(\d{3})$', codigo)
                            if match:
                                mes_encontrado = match.group(1)
                                tipo_encontrado = match.group(2)
                                secuencia_str = match.group(3)

                                if (mes_encontrado == mes or mes_encontrado == patron_mes) and tipo_encontrado == tipo:
                                    try:
                                        secuencia = int(secuencia_str)
                                        ultima_secuencia = max(ultima_secuencia, secuencia)
                                        print(f"Secuencia encontrada: {secuencia}")
                                    except ValueError:
                                        print(f"Error al convertir secuencia: {secuencia_str}")

                    data['secuencia'] = ultima_secuencia
                    print(f"Secuencia devuelta: {ultima_secuencia}")

                except Exception as e:
                    import traceback
                    print(f"Error al buscar secuencia: {str(e)}")
                    print(traceback.format_exc())
                    data['secuencia'] = 0
                    data['error'] = str(e)

            elif action == 'create':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = EncabezadoCuentasPlanCuenta()
                    encabezado.codigo = request.POST['codigo']
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.empresa_id = request.POST['empresa']
                    encabezado.proveedor_id = request.POST['proveedor']
                    encabezado.save()

                    for i in items:
                        try:
                            cuenta_id = int(i['id'])
                            if not PlanCuenta.objects.filter(id=cuenta_id).exists():
                                raise ValueError(f"La cuenta con ID {cuenta_id} no existe en el plan de cuentas")

                            cuerpo = DetalleCuentasPlanCuenta()
                            cuerpo.encabezadocuentaplan_id = encabezado.pk
                            cuerpo.cuenta_id = cuenta_id
                            cuerpo.detalle = i['detalle']
                            cuerpo.debe = Decimal(str(i.get('debe', '0')).replace(',', '.')) if i.get(
                                'debe') else Decimal('0.00')
                            cuerpo.haber = Decimal(str(i.get('haber', '0')).replace(',', '.')) if i.get(
                                'haber') else Decimal('0.00')
                            cuerpo.save()
                        except Exception as e:
                            transaction.set_rollback(True)
                            data['error'] = f"Error al guardar el detalle: {str(e)}"
                            print(f"Error al guardar detalle: {str(e)}")
                            return JsonResponse(data, safe=False)
            else:
                data['error'] = 'Ha ocurrido un error'

        except Exception as e:
            import traceback
            print("Error en la vista:")
            print(traceback.format_exc())
            data['error'] = f'Error: {str(e)}'

        return JsonResponse(data, safe=False)

    def search_plan_improved(self, request):
        """Función mejorada para búsqueda del plan de cuentas"""
        try:
            empresa = request.POST.get('empresa', 'BIO')
            page = int(request.POST.get('page', 1))
            page_size = int(request.POST.get('page_size', 500))
            search_term = request.POST.get('search', '').strip()
            search_type = request.POST.get('search_type', 'all')  # 'all', 'exact', 'partial'
            print(f'Búsqueda: página={page}, tamaño={page_size}, término="{search_term}", tipo={search_type}')
            ids_exclude = []
            try:
                ids_exclude = json.loads(request.POST.get('ids', '[]'))
            except:
                ids_exclude = []
            queryset = PlanCuenta.objects.filter(empresa__siglas__exact=empresa).exclude(id__in=ids_exclude)
            if search_term:
                if search_type == 'exact':
                    queryset = queryset.filter(codigo__exact=search_term)
                elif search_type == 'partial':
                    queryset = queryset.filter(
                        Q(codigo__icontains=search_term) |
                        Q(nombre__icontains=search_term)
                    )
                else:
                    queryset = queryset.filter(
                        Q(codigo__icontains=search_term) |
                        Q(nombre__icontains=search_term) |
                        Q(tipo_cuenta__icontains=search_term)
                    )
            queryset = queryset.order_by('codigo', 'nombre')
            total_count = queryset.count()
            print(f'Total de registros encontrados: {total_count}')
            paginator = Paginator(queryset, page_size)
            try:
                page_obj = paginator.get_page(page)
            except:
                page_obj = paginator.get_page(1)
            data = []
            for item in page_obj:
                item_data = item.toJSON()
                item_data['detalle'] = ""
                data.append(item_data)
            response_data = {
                'data': data,
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_records': total_count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                    'page_size': page_size
                },
                'search_info': {
                    'term': search_term,
                    'type': search_type,
                    'found_count': total_count
                }
            }
            print(f'Enviando {len(data)} registros de {total_count} totales')
            return JsonResponse(response_data, safe=False)

        except Exception as e:
            print(f'Error en search_plan_improved: {str(e)}')
            import traceback
            print(traceback.format_exc())

            return JsonResponse({
                'error': f'Error al cargar datos: {str(e)}',
                'data': [],
                'pagination': {
                    'current_page': 1,
                    'total_pages': 0,
                    'total_records': 0,
                    'has_next': False,
                    'has_previous': False,
                    'page_size': page_size
                }
            }, status=500)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'create'
        planCuenta = PlanCuenta.objects.filter(parentId=None, empresa__siglas__exact='BIO')
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.filter(empresa__siglas__exact='BIO')
        context['planCuenta2'] = planCuenta2
        context['empresa'] = 'BIO'
        context['det'] = []
        try:
            empresa_bio = Empresa.objects.get(siglas='BIO')
            form = self.get_form()
            form.fields['empresa'].initial = empresa_bio.id
            context['form'] = form
        except Exception as e:
            print(f"Error al preseleccionar empresa BIO: {e}")
        return context


class editarTransaccionPlanBIOView(UpdateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crearBIO.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta_BIO')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']

            if action == 'search_plan':
                # Usar la misma función mejorada de búsqueda
                return self.search_plan_improved(request)

            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST.get('ids', '[]'))
                term = request.POST.get('term', '').strip()

                data.append({'codigo': term, 'text': term, 'id': None})

                plan_detail = PlanCuenta.objects.filter(
                    Q(nombre__icontains=term) | Q(codigo__icontains=term),
                    empresa__siglas__exact='BIO'
                ).exclude(id__in=ids_exclude).order_by('codigo')[:50]

                for i in plan_detail:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = f"{i.codigo} - {i.nombre}"
                    item['id'] = int(i.id) if i.id else None
                    data.append(item)

            elif action == 'obtener_ultima_secuencia':
                # En modo edición, NUNCA generar nuevo código
                encabezado_actual = self.get_object()
                codigo_original = str(encabezado_actual.codigo) if encabezado_actual.codigo else ""

                print(f"MODO EDICIÓN - Código original: {codigo_original}")
                print("IMPORTANTE: NO se debe generar nuevo código en edición")

                data = {
                    'codigo_original': codigo_original,
                    'mantener_codigo': True,
                    'es_edicion': True,
                    'secuencia': 0,
                    'mensaje': 'Código original mantenido'
                }

                print(f"Devolviendo código original sin cambios: {codigo_original}")

            elif action == 'edit':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = self.get_object()

                    # NUNCA cambiar el código en edición
                    codigo_original = str(encabezado.codigo) if encabezado.codigo else ""
                    print(f"Manteniendo código original: {codigo_original}")

                    # Actualizar otros campos (NO el código)
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.empresa_id = request.POST['empresa']
                    encabezado.proveedor_id = request.POST['proveedor']
                    encabezado.save()

                    # Eliminar detalles existentes y crear nuevos
                    encabezado.detallecuentasplancuenta_set.all().delete()

                    for i in items:
                        cuerpo = DetalleCuentasPlanCuenta()
                        cuerpo.encabezadocuentaplan_id = encabezado.pk
                        cuerpo.cuenta_id = int(i['id'])
                        cuerpo.detalle = i['detalle']

                        try:
                            debe_value = str(i.get('debe', '0')).replace(',', '.')
                            haber_value = str(i.get('haber', '0')).replace(',', '.')

                            cuerpo.debe = float(debe_value) if debe_value else 0
                            cuerpo.haber = float(haber_value) if haber_value else 0
                        except (ValueError, TypeError) as e:
                            print(
                                f"Error al convertir valores: debe={i.get('debe')}, haber={i.get('haber')}, error={e}")
                            cuerpo.debe = 0
                            cuerpo.haber = 0

                        cuerpo.save()

            else:
                data['error'] = 'Ha ocurrido un error'

        except Exception as e:
            data['error'] = f'Error: {str(e)}'
            print(f"Error en vista editar: {str(e)}")
            import traceback
            print(traceback.format_exc())

        return JsonResponse(data, safe=False)

    def search_plan_improved(self, request):
        """Función mejorada para búsqueda del plan de cuentas (compartida con crear)"""
        try:
            empresa = request.POST.get('empresa', 'BIO')
            page = int(request.POST.get('page', 1))
            page_size = int(request.POST.get('page_size', 500))
            search_term = request.POST.get('search', '').strip()
            search_type = request.POST.get('search_type', 'all')

            print(f'Búsqueda (edición): página={page}, término="{search_term}", tipo={search_type}')

            ids_exclude = []
            try:
                ids_exclude = json.loads(request.POST.get('ids', '[]'))
            except:
                ids_exclude = []

            queryset = PlanCuenta.objects.filter(
                empresa__siglas__exact=empresa
            ).exclude(id__in=ids_exclude)

            if search_term:
                if search_type == 'exact':
                    queryset = queryset.filter(codigo__exact=search_term)
                elif search_type == 'partial':
                    queryset = queryset.filter(
                        Q(codigo__icontains=search_term) |
                        Q(nombre__icontains=search_term)
                    )
                else:
                    queryset = queryset.filter(
                        Q(codigo__icontains=search_term) |
                        Q(nombre__icontains=search_term) |
                        Q(tipo_cuenta__icontains=search_term)
                    )

            queryset = queryset.order_by('codigo', 'nombre')
            total_count = queryset.count()

            paginator = Paginator(queryset, page_size)
            try:
                page_obj = paginator.get_page(page)
            except:
                page_obj = paginator.get_page(1)

            data = []
            for item in page_obj:
                item_data = item.toJSON()
                item_data['detalle'] = ""
                data.append(item_data)

            response_data = {
                'data': data,
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_records': total_count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                    'page_size': page_size
                },
                'search_info': {
                    'term': search_term,
                    'type': search_type,
                    'found_count': total_count
                }
            }

            return JsonResponse(response_data, safe=False)

        except Exception as e:
            print(f'Error en search_plan_improved (edición): {str(e)}')
            return JsonResponse({
                'error': f'Error al cargar datos: {str(e)}',
                'data': [],
                'pagination': {
                    'current_page': 1,
                    'total_pages': 0,
                    'total_records': 0,
                    'has_next': False,
                    'has_previous': False,
                    'page_size': page_size
                }
            }, status=500)

    def get_detalle(self):
        data = []
        for i in DetalleCuentasPlanCuenta.objects.filter(encabezadocuentaplan_id=self.kwargs['pk']):
            item = i.cuenta.toJSON()
            item['detalle'] = i.detalle
            item['debe'] = format(i.debe, '.2f')
            item['haber'] = format(i.haber, '.2f')
            data.append(item)
        return json.dumps(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Edición de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'edit'

        # Filtrar solo cuentas BIO
        planCuenta = PlanCuenta.objects.filter(parentId=None, empresa__siglas__exact='BIO')
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.filter(empresa__siglas__exact='BIO')
        context['planCuenta2'] = planCuenta2
        context['det'] = self.get_detalle()

        # Pasar el código original al contexto
        context['codigo_original'] = str(self.object.codigo) if self.object.codigo else ""

        return context


class editarTransaccionPlanView(UpdateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_plan':
                data = []
                queryset = PlanCuenta.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                if len(ids_exclude):
                    queryset = queryset.filter().exclude(id__in=ids_exclude).order_by('codigo')
                for i in queryset.order_by('id'):
                    item = i.toJSON()
                    item['detalle'] = ""
                    data.append(item)
            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST['ids'])
                term = request.POST['term'].strip()
                data.append({'codigo': term, 'text': term})
                plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
                for i in plan_detail[0:50]:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = i.nombre
                    data.append(item)
            elif action == 'edit':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = self.get_object()
                    encabezado.codigo = request.POST['codigo']
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.empresa_id = request.POST['empresa']
                    encabezado.save()
                    for s in encabezado.detallecuentasplancuenta_set.all():
                        print('s del recorredor')
                        print(s)
                    # encabezado.detallecuentasplancuenta_set.all().delete()
                    for i in items:
                        cuerpo = DetalleCuentasPlanCuenta()
                        cuerpo.encabezadocuentaplan_id = encabezado.pk
                        cuerpo.cuenta_id = int(i['id'])
                        cuerpo.detalle = i['detalle']
                        cuerpo.debe = int(i['debe']) if i.get('debe') else 0
                        cuerpo.haber = int(i['haber']) if i.get('haber') else 0
                        cuerpo.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_detalle(self):
        data = []
        print('self.get_object().id')
        print(self.get_object().id)
        print("self.kwargs['pk']")
        print(self.kwargs['pk'])
        for i in DetalleCuentasPlanCuenta.objects.filter(encabezadocuentaplan_id=self.kwargs['pk']):
            item = i.cuenta.toJSON()
            item['detalle'] = i.detalle
            item['debe'] = format(i.debe, '.2f')
            item['haber'] = format(i.haber, '.2f')
            data.append(item)
        return json.dumps(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.all()
        context['planCuenta2'] = planCuenta2
        context['det'] = self.get_detalle()
        return context




#
# class eliminarTransaccionPlanView(DeleteView):
#     model = InvoiceStock
#     template_name = 'app_factura_detalle/factura_detalle_eliminar.html'
#     success_url = reverse_lazy('app_factura:listar_factura')
#     url_redirect = success_url
#
#     @method_decorator(csrf_exempt)
#     def dispatch(self, request, *args, **kwargs):
#         self.object = self.get_object()
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             self.object.delete()
#         except Exception as e:
#             data['error'] = str(e)
#         return JsonResponse(data)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['title'] = 'Eliminación de una Factura'
#         context['entity'] = 'Factura'
#         context['list_url'] = reverse_lazy('app_factura:listar_factura')
#         return context
#
# class reporteTransaccionPlanView(View):
#
#     def get(self, request, *args, **kwargs):
#         try:
#             template = get_template('app_reportes/factura_reporte.html')
#             detalle = Producto_Stock.objects.filter(invoice_stock_id=self.kwargs['pk'])
#
#             empresa = ''
#             # fecha_ingreso = ''
#             # numero_guia = ''
#             # number = 0
#             # proveedor = ''
#
#             if detalle:
#                 empresa = detalle[0].producto_empresa.nombre_empresa
#                 # fecha_ingreso = detalle[0].invoice_stock.fecha_ingreso
#                 # numero_guia = detalle[0].invoice_stock.numero_guia
#                 # number = detalle[0].invoice_stock.id
#                 # proveedor = detalle[0].invoice_stock.proveedor
#
#             context = {
#                 'sale': InvoiceStock.objects.get(pk=self.kwargs['pk']),
#                 'comp': {'name': 'INDUSTRIA PESQUERA', 'address': 'MACHALA - EL ORO - ECUADOR',
#                          'numero': '(072) 920 371', 'comprobante': 'COMPROBANTE DE INGRESOs DE PRODUCTOS'},
#                 'icon': '{}{}'.format(settings.MEDIA_URL, 'logo.png'),
#                 'empresa': empresa,
#                 # 'detalle_fac': detalle,
#                 # 'fecha_ingreso': fecha_ingreso,
#                 # 'numero_guia': numero_guia,
#                 # 'number': number,
#                 # 'proveedor': proveedor,
#                 # 'novedad': Producto_Stock.objects.filter(invoice_stock_id=self.kwargs['pk'])
#             }
#             print('context')
#             print(context)
#             html = template.render(context)
#             css_url = os.path.join(settings.BASE_DIR, 'static/lib/bootstrap-4.4.1-dist/css/bootstrap.min.css')
#             pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
#             return HttpResponse(pdf, content_type='application/pdf')
#         except:
#             pass
#         return HttpResponseRedirect(reverse_lazy('app_factura:listar_factura'))


# ATS - ANEXO TRANSACCIONAL


# LISTAR ANEXO-TRANSACCIONAL
class listarAnexoTransaccionalView(ListView):
    model = EncabezadoCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_listar.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in EncabezadoCuentasPlanCuenta.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Transaccion de Plan de Cuentas'
        context['title'] = 'Transaccion de Plan de Cuentas'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


# CREAR ANEXO-TRANSACCIONAL
class crearAnexoTransaccionalView(CreateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_plan':
                data = []
                queryset = PlanCuenta.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                if len(ids_exclude):
                    queryset = queryset.filter().exclude(id__in=ids_exclude)
                for i in queryset:
                    item = i.toJSON()
                    item['detalle'] = ""
                    data.append(item)

            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST['ids'])
                term = request.POST['term'].strip()
                data.append({'codigo': term, 'text': term})
                plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
                for i in plan_detail[0:50]:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = i.nombre
                    data.append(item)

            elif action == 'create':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = EncabezadoCuentasPlanCuenta()
                    encabezado.codigo = request.POST['codigo']
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.save()
                    for i in items:
                        cuerpo = DetalleCuentasPlanCuenta()
                        cuerpo.encabezadocuentaplan_id = encabezado.pk
                        cuerpo.cuenta_id = int(i['id'])
                        cuerpo.detalle = i['detalle']
                        cuerpo.debe = int(i['debe']) if i.get('debe') else 0
                        cuerpo.haber = int(i['haber']) if i.get('haber') else 0
                        cuerpo.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'create'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.all()
        context['planCuenta2'] = planCuenta2
        context['det'] = []
        return context


class editarAnexoTransaccionalView(UpdateView):
    model = EncabezadoCuentasPlanCuenta
    form_class = EncabezadoCuentasPlanCuentaForm
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/transaccionPlan_crear.html'
    success_url = reverse_lazy('app_planCuentas:listar_planCuenta')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_plan':
                data = []
                queryset = PlanCuenta.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                if len(ids_exclude):
                    queryset = queryset.filter().exclude(id__in=ids_exclude).order_by('codigo')
                for i in queryset.order_by('id'):
                    item = i.toJSON()
                    item['detalle'] = ""
                    data.append(item)
            elif action == 'search_autocomplete':
                data = []
                ids_exclude = json.loads(request.POST['ids'])
                term = request.POST['term'].strip()
                data.append({'codigo': term, 'text': term})
                plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
                for i in plan_detail[0:50]:
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = i.nombre
                    data.append(item)
            elif action == 'edit':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    encabezado = self.get_object()
                    encabezado.codigo = request.POST['codigo']
                    encabezado.tip_cuenta = request.POST['tip_cuenta']
                    encabezado.fecha = request.POST['fecha']
                    encabezado.comprobante = request.POST['comprobante']
                    encabezado.descripcion = request.POST['descripcion']
                    encabezado.direccion = request.POST['direccion']
                    encabezado.save()
                    for s in encabezado.detallecuentasplancuenta_set.all():
                        print('s del recorredor')
                        print(s)
                    # encabezado.detallecuentasplancuenta_set.all().delete()
                    for i in items:
                        cuerpo = DetalleCuentasPlanCuenta()
                        cuerpo.encabezadocuentaplan_id = encabezado.pk
                        cuerpo.cuenta_id = int(i['id'])
                        cuerpo.detalle = i['detalle']
                        cuerpo.debe = int(i['debe']) if i.get('debe') else 0
                        cuerpo.haber = int(i['haber']) if i.get('haber') else 0
                        cuerpo.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_detalle(self):
        data = []
        # print('self.get_object().id')
        # print(self.get_object().id)
        # print("self.kwargs['pk']")
        # print(self.kwargs['pk'])
        for i in DetalleCuentasPlanCuenta.objects.filter(encabezadocuentaplan_id=self.kwargs['pk']):
            item = i.cuenta.toJSON()
            item['detalle'] = i.detalle
            item['debe'] = format(i.debe, '.2f')
            item['haber'] = format(i.haber, '.2f')
            data.append(item)
        return json.dumps(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Formulario de Registro de Transacción'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        planCuenta = PlanCuenta.objects.filter(parentId=None)
        context['planCuenta'] = planCuenta
        planCuenta2 = PlanCuenta.objects.all()
        context['planCuenta2'] = planCuenta2
        context['det'] = self.get_detalle()
        return context


# LIBRO MAYOR
class listarMayorPlanView(ListView):
    model = DetalleCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/mayorizacion/mayorizacionPlan_listar.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                empresa = request.POST['empresa']
                print('empresa')
                print(empresa)
                detallecuenta = DetalleCuentasPlanCuenta.objects.filter(
                    cuenta__empresa__siglas__exact=empresa).order_by('cuenta__codigo')
                for i in detallecuenta:
                    data.append(i.toJSON())

            elif action == 'searchdataplan':
                print('llego a search data plan')
                data = []
                for i in PlanCuenta.objects.filter(empresa__siglas__exact='PSM').order_by('codigo'):
                    item = i.toJSON()
                    item['codigo'] = i.codigo
                    item['text'] = i.get_name()
                    data.append(item)
            # elif action == 'search_plan':
            #     print('LLEGO A SEARCH PLAN')
            #     data = []
            #     term = request.POST['term'].strip()
            #     queryset = PlanCuenta.objects.filter(codigo__icontains=term)
            #     for i in queryset:
            #         item = i.toJSON()
            #         item['codigo'] = i.codigo
            #         item['text'] = i.nombre
            #         data.append(item)
            #
            # elif action == 'search_autocomplete':
            #     print('LLEGO A SEARCH AUTOCOMPLETE')
            #     data = []
            #     term = request.POST['term']
            #     print('se extrajo parametro de term')
            #     print(term)
            #     data.append({'codigo': term, 'text': term})
            #     plan_detail = PlanCuenta.objects.filter(Q(nombre__icontains=term))[0:50]
            #     for i in plan_detail:
            #         item = i.toJSON()
            #         data.append(item)
            #
            # elif action == 'search_autocomplete':
            #     data = []
            #     ids_exclude = json.loads(request.POST['ids'])
            #     term = request.POST['term'].strip()
            #     data.append({'codigo': term, 'text': term})
            #     plan_detail = PlanCuenta.objects.filter(nombre__icontains=term).exclude(id__in=ids_exclude)
            #     for i in plan_detail[0:50]:
            #         item = i.toJSON()
            #         item['codigo'] = i.codigo
            #         item['text'] = i.nombre
            #         data.append(item)

            elif action == 'search_report':
                data = []
                start_date = request.POST.get('start_date', '')
                end_date = request.POST.get('end_date', '')
                desde_rang = request.POST.get('desde_rang', '')
                hasta_rang = request.POST.get('hasta_rang', '')
                empresa = request.POST['empresa']
                print('empresa')
                print(empresa)
                search = DetalleCuentasPlanCuenta.objects.all()
                if len(start_date) and len(end_date):
                    search = search.filter(
                        encabezadocuentaplan__fecha__range=[start_date, end_date],
                        cuenta__codigo__range=[desde_rang, hasta_rang],
                        cuenta__empresa__siglas__exact=empresa,
                    )
                for i in search:
                    data.append(i.toJSON())

            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Libro Mayor Empresa PSM'
        context['title'] = 'Libro Mayor Empresa PSM'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


# LIBRO MAYOR
class listarMayorPlanViewBIO(ListView):

    model = DetalleCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/transaccion_Plan/mayorizacion/mayorizacionPlan_listarBIO.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        try:

            action = request.POST.get('action')

            # ==========================================================
            # DATATABLE PRINCIPAL
            # ==========================================================
            if action == 'searchdata':

                draw = int(request.POST.get('draw', 1))
                start = int(request.POST.get('start', 0))
                length = int(request.POST.get('length', 25))
                search_value = request.POST.get('search[value]', '')
                empresa = request.POST.get('empresa', 'BIO')

                queryset = DetalleCuentasPlanCuenta.objects.filter(
                    cuenta__empresa__siglas=empresa
                ).select_related(
                    'cuenta',
                    'encabezadocuentaplan'
                )

                # ======================================================
                # BUSCADOR
                # ======================================================

                if search_value:
                    queryset = queryset.filter(
                        Q(cuenta__codigo__icontains=search_value) |
                        Q(cuenta__nombre__icontains=search_value) |
                        Q(descripcion__icontains=search_value)
                    )

                records_total = queryset.count()

                # ======================================================
                # ORDEN
                # ======================================================

                queryset = queryset.order_by(
                    'cuenta__codigo',
                    'id'
                )[start:start + length]

                # ======================================================
                # DATA
                # ======================================================

                data = []

                saldo_acumulado = 0
                cuenta_actual = None

                for i in queryset:

                    codigo_cuenta = i.cuenta.codigo if i.cuenta else ''

                    debe = float(i.debe) if i.debe else 0
                    haber = float(i.haber) if i.haber else 0

                    # ==================================================
                    # REINICIAR SALDO SI CAMBIA LA CUENTA
                    # ==================================================

                    if cuenta_actual != codigo_cuenta:
                        saldo_acumulado = 0
                        cuenta_actual = codigo_cuenta

                    # ==================================================
                    # CALCULAR SALDO
                    # ==================================================

                    saldo_acumulado += (debe - haber)

                    data.append({
                        'codigo': codigo_cuenta,
                        'nombre': i.cuenta.nombre if i.cuenta else '',
                        'descripcion': getattr(i, 'descripcion', ''),
                        'fecha': str(i.encabezadocuentaplan.fecha) if i.encabezadocuentaplan else '',
                        'transaccion': getattr(i, 'transaccion', ''),
                        'asiento': getattr(i, 'asiento', ''),
                        'debe': round(debe, 2),
                        'haber': round(haber, 2),
                        'saldo': round(saldo_acumulado, 2),
                    })

                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': records_total,
                    'recordsFiltered': records_total,
                    'data': data
                })

            # ==========================================================
            # SELECT PLAN CUENTAS
            # ==========================================================
            elif action == 'searchdataplan':

                data = []

                queryset = PlanCuenta.objects.filter(
                    empresa__siglas='BIO'
                ).only(
                    'id',
                    'codigo',
                    'nombre'
                ).order_by('codigo')

                for i in queryset:

                    data.append({
                        'id': i.id,
                        'codigo': i.codigo,
                        'text': f'{i.codigo} / {i.nombre}'
                    })

                return JsonResponse(data, safe=False)

            # ==========================================================
            # REPORTE POR RANGO
            # ==========================================================
            elif action == 'search_report':

                draw = int(request.POST.get('draw', 1))
                start = int(request.POST.get('start', 0))
                length = int(request.POST.get('length', 25))

                start_date = request.POST.get('start_date', '')
                end_date = request.POST.get('end_date', '')
                desde_rang = request.POST.get('desde_rang', '')
                hasta_rang = request.POST.get('hasta_rang', '')
                empresa = request.POST.get('empresa', 'BIO')

                queryset = DetalleCuentasPlanCuenta.objects.filter(
                    cuenta__empresa__siglas=empresa
                ).select_related(
                    'cuenta',
                    'encabezadocuentaplan'
                )

                # ======================================================
                # FILTRO FECHAS
                # ======================================================

                if start_date and end_date:

                    queryset = queryset.filter(
                        encabezadocuentaplan__fecha__range=[
                            start_date,
                            end_date
                        ]
                    )

                # ======================================================
                # FILTRO RANGO CUENTAS
                # ======================================================

                if desde_rang and hasta_rang:

                    queryset = queryset.filter(
                        cuenta__codigo__range=[
                            desde_rang,
                            hasta_rang
                        ]
                    )

                records_total = queryset.count()

                # ======================================================
                # ORDEN
                # ======================================================

                queryset = queryset.order_by(
                    'cuenta__codigo',
                    'id'
                )[start:start + length]

                # ======================================================
                # DATA
                # ======================================================

                data = []

                saldo_acumulado = 0
                cuenta_actual = None

                for i in queryset:

                    codigo_cuenta = i.cuenta.codigo if i.cuenta else ''

                    debe = float(i.debe) if i.debe else 0
                    haber = float(i.haber) if i.haber else 0

                    # ==================================================
                    # REINICIAR SALDO SI CAMBIA LA CUENTA
                    # ==================================================

                    if cuenta_actual != codigo_cuenta:
                        saldo_acumulado = 0
                        cuenta_actual = codigo_cuenta

                    # ==================================================
                    # CALCULAR SALDO
                    # ==================================================

                    saldo_acumulado += (debe - haber)

                    data.append({
                        'codigo': codigo_cuenta,
                        'nombre': i.cuenta.nombre if i.cuenta else '',
                        'descripcion': getattr(i, 'descripcion', ''),
                        'fecha': str(i.encabezadocuentaplan.fecha) if i.encabezadocuentaplan else '',
                        'transaccion': getattr(i, 'transaccion', ''),
                        'asiento': getattr(i, 'asiento', ''),
                        'debe': round(debe, 2),
                        'haber': round(haber, 2),
                        'saldo': round(saldo_acumulado, 2),
                    })

                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': records_total,
                    'recordsFiltered': records_total,
                    'data': data
                })

            # ==========================================================
            # ERROR
            # ==========================================================

            return JsonResponse({
                'error': 'No existe la acción solicitada'
            })

        except Exception as e:

            return JsonResponse({
                'error': str(e)
            })

    # ==============================================================
    # CONTEXT
    # ==============================================================

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['nombre'] = 'Libro Mayor Empresa BIO'
        context['title'] = 'Libro Mayor Empresa BIO'
        context['list_url'] = reverse_lazy(
            'app_planCuentas:listar_transaccionPlan'
        )

        return context


# BALANCE PLAN
# class listarBalancePlanView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         print('request.POST')
#         print(request.POST)
#         data = {}
#         try:
#             action = request.POST['action']
#             if action == 'searchdata_psm':
#                 data = []
#                 detallecuenta = DetalleCuentasPlanCuenta.objects.filter(cuenta__empresa__siglas__exact='PSM').order_by('cuenta__codigo')
#                 for i in detallecuenta:
#                     data.append(i.toJSON())
#
#             elif action == 'searchdataplan':
#                 print('llego a search data LIBRO DETALLE plan')
#                 data = []
#                 for i in PlanCuenta.objects.all().order_by('codigo'):
#                     item = i.toJSON()
#                     item['codigo'] = i.codigo
#                     item['text'] = i.get_name()
#                     data.append(item)
#
#             elif action == 'searchdatahierarchy':
#                 print('Generando jerarquía del plan de cuentas')
#
#                 # Construir jerarquía
#                 detallecuenta = DetalleCuentasPlanCuenta.objects.select_related('cuenta').all().order_by('cuenta__codigo')
#                 hierarchy = defaultdict(lambda: {"children": [], "cuenta": None})
#
#                 for item in detallecuenta:
#                     cuenta = item.cuenta
#                     codigo = cuenta.codigo
#                     parent_codigo = ".".join(codigo.split(".")[:-1]) if "." in codigo else None
#
#                     # Crear nodo
#                     node = {
#                         "codigo": cuenta.codigo,
#                         "nombre": cuenta.nombre,
#                         "debe": item.debe,
#                         "haber": item.haber,
#                         "saldo": item.debe - item.haber,  # Calcula el saldo
#                     }
#                     hierarchy[codigo]["cuenta"] = node
#
#                     # Agregar nodo al padre
#                     if parent_codigo:
#                         hierarchy[parent_codigo]["children"].append(hierarchy[codigo])
#
#                 # Filtrar nodos raíz
#                 data = [node for key, node in hierarchy.items() if "." not in key]
#
#             elif action == 'search_report':
#                 print('llego a search data RANGUE LIBRO DETALLE plan')
#                 data = []
#                 start_date = request.POST.get('start_date', '')
#                 end_date = request.POST.get('end_date', '')
#                 desde_rang = request.POST.get('desde_rang', '')
#                 hasta_rang = request.POST.get('hasta_rang', '')
#                 search = DetalleCuentasPlanCuenta.objects.all()
#                 if len(start_date) and len(end_date):
#                     search = search.filter(
#                         encabezadocuentaplan__fecha__range=[start_date, end_date],
#                         cuenta__codigo__range=[desde_rang, hasta_rang],
#                     )
#                 for i in search:
#                     data.append(i.toJSON())
#
#             else:
#                 data['error'] = 'Ha ocurrido un error'
#         except Exception as e:
#             data['error'] = str(e)
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context


# ESTE ES EL QUE CASI ME GUSTO
# class listarBalancePlanView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
#         saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__lt=fecha_inicio
#         ).aggregate(
#             debe=Coalesce(Sum('debe'), Decimal('0')),
#             haber=Coalesce(Sum('haber'), Decimal('0'))
#         )
#
#         saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin]
#         ).aggregate(
#             debe=Coalesce(Sum('debe'), Decimal('0')),
#             haber=Coalesce(Sum('haber'), Decimal('0'))
#         )
#
#         saldo_actual_debe = saldo_anterior['debe'] + saldo_mes['debe']
#         saldo_actual_haber = saldo_anterior['haber'] + saldo_mes['haber']
#
#         return {
#             'saldo_anterior_debe': float(saldo_anterior['debe']),
#             'saldo_anterior_haber': float(saldo_anterior['haber']),
#             'debe': float(saldo_mes['debe']),
#             'haber': float(saldo_mes['haber']),
#             'saldo_actual_debe': float(saldo_actual_debe),
#             'saldo_actual_haber': float(saldo_actual_haber)
#         }
#
#     def get_parent_codes(self, codigo):
#         """Obtiene los códigos padre de una cuenta basado en su código"""
#         parent_codes = []
#         code = codigo
#         while len(code) > 1:
#             # Remove last segment of the code
#             code = code[:-2] if len(code) > 2 else code[:-1]
#             if code:
#                 parent_codes.append(code)
#         return parent_codes
#
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         data = []
#         try:
#             action = request.POST.get('action')
#             if action == 'searchdata_psm':
#                 fecha_inicio = request.POST.get('fecha_inicio')
#                 fecha_fin = request.POST.get('fecha_fin')
#
#                 if not fecha_inicio or not fecha_fin:
#                     return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)
#
#                 try:
#                     fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
#                     fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
#                 except ValueError:
#                     return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
#
#                 # Obtener todas las cuentas ordenadas por código
#                 cuentas = PlanCuenta.objects.filter(
#                     empresa__siglas__exact='PSM'
#                 ).order_by('codigo')
#
#                 # Crear un diccionario para almacenar las cuentas por código
#                 cuentas_dict = {}
#
#                 for cuenta in cuentas:
#                     saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)
#
#                     # Solo incluir cuentas con movimientos o que sean cuentas mayores
#                     if (saldos['saldo_anterior_debe'] != 0 or
#                             saldos['saldo_anterior_haber'] != 0 or
#                             saldos['debe'] != 0 or
#                             saldos['haber'] != 0 or
#                             len(cuenta.codigo) <= 3):  # Cuentas mayores
#
#                         item = {
#                             'id': cuenta.id,
#                             'codigo': cuenta.codigo,
#                             'nombre': cuenta.nombre,
#                             'nivel': len(self.get_parent_codes(cuenta.codigo)) + 1,
#                             'parent_codigo': self.get_parent_codes(cuenta.codigo)[0] if self.get_parent_codes(
#                                 cuenta.codigo) else None,
#                             **saldos
#                         }
#                         data.append(item)
#                         cuentas_dict[cuenta.codigo] = item
#
#             else:
#                 return JsonResponse({'error': 'Acción no reconocida'}, status=400)
#
#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)
#
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context


class listarBalancePlanView(ListView):
    model = DetalleCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        """
        Calcula los saldos de una cuenta para un período específico

        Saldo anterior: Transacciones dentro del rango de fechas seleccionado
        Saldo mes: Transacciones del mes actual
        Saldo actual: Suma de ambos
        """
        # Obtener la fecha actual para determinar el mes en curso
        hoy = datetime.now().date()
        primer_dia_mes_actual = datetime(hoy.year, hoy.month, 1).date()

        # Calcular el último día del mes actual
        if hoy.month == 12:
            ultimo_dia_mes_actual = datetime(hoy.year + 1, 1, 1).date() - timedelta(days=1)
        else:
            ultimo_dia_mes_actual = datetime(hoy.year, hoy.month + 1, 1).date() - timedelta(days=1)

        # Saldo anterior: todas las transacciones dentro del rango de fechas seleccionado
        saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
            cuenta=cuenta,
            encabezadocuentaplan__fecha__gte=fecha_inicio,
            encabezadocuentaplan__fecha__lte=fecha_fin
        ).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        # Saldo del mes: transacciones del mes actual
        saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
            cuenta=cuenta,
            encabezadocuentaplan__fecha__gte=primer_dia_mes_actual,
            encabezadocuentaplan__fecha__lte=ultimo_dia_mes_actual
        ).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        # Depuración
        logger.debug(f"Cuenta: {cuenta.codigo}, Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")
        logger.debug(f"Mes actual: {primer_dia_mes_actual} a {ultimo_dia_mes_actual}")
        logger.debug(f"Cuenta: {cuenta.codigo}, Saldo anterior: {saldo_anterior}, Saldo mes: {saldo_mes}")

        # Para cuentas padre, incluir saldos de subcuentas
        subcuentas = PlanCuenta.objects.filter(parentId=cuenta)
        if subcuentas.exists():
            for subcuenta in subcuentas:
                sub_saldos = self.get_saldos_cuenta(subcuenta, fecha_inicio, fecha_fin)
                saldo_anterior['debe'] += Decimal(str(sub_saldos['debe_ant']))
                saldo_anterior['haber'] += Decimal(str(sub_saldos['haber_ant']))
                saldo_mes['debe'] += Decimal(str(sub_saldos['debe_mes']))
                saldo_mes['haber'] += Decimal(str(sub_saldos['haber_mes']))

        debe_actual = saldo_anterior['debe'] + saldo_mes['debe']
        haber_actual = saldo_anterior['haber'] + saldo_mes['haber']

        return {
            'debe_ant': float(saldo_anterior['debe']),
            'haber_ant': float(saldo_anterior['haber']),
            'debe_mes': float(saldo_mes['debe']),
            'haber_mes': float(saldo_mes['haber']),
            'debe_act': float(debe_actual),
            'haber_act': float(haber_actual)
        }

    def get_account_type(self, cuenta):
        """Determina el tipo de cuenta basado en su código y naturaleza"""
        if len(cuenta.codigo) == 1:  # Cuenta mayor
            return 'G'
        elif len(cuenta.codigo) >= 9:  # Si tiene 9 o más caracteres es cuenta de detalle
            return 'D'
        else:
            return 'G'

    def get_nivel(self, codigo):
        """Determina el nivel de la cuenta basado en su código"""
        partes = codigo.split('.')
        return len(partes)

    def format_nombre_cuenta(self, nombre, codigo):
        """
        Formatea el nombre de la cuenta con la indentación correcta basada en el código
        """
        # Calculamos el nivel basado en la cantidad de segmentos en el código
        nivel = len(codigo.split('.'))

        # Aplicamos la indentación (2 espacios por nivel, excepto nivel 1)
        if nivel == 1:
            return nombre
        else:
            indentacion = ' ' * ((nivel - 1) * 2)  # 2 espacios por nivel
            return f"{indentacion}{nombre}"

    def get_parent_code(self, codigo):
        """Obtiene el código del padre de una cuenta"""
        parts = codigo.split('.')
        if len(parts) > 1:
            return '.'.join(parts[:-1])
        return None

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            action = request.POST.get('action')
            if action == 'searchdata_psm':
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin = request.POST.get('fecha_fin')
                cuenta_inicio = request.POST.get('cuenta_inicio', '1')
                cuenta_fin = request.POST.get('cuenta_fin', '9')

                if not fecha_inicio or not fecha_fin:
                    return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)

                try:
                    # Convertir las fechas asegurando que estén en el formato correcto
                    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                    # Validar que fecha_fin no sea menor que fecha_inicio
                    if fecha_fin < fecha_inicio:
                        return JsonResponse({
                            'error': 'La fecha final no puede ser menor que la fecha inicial'
                        }, status=400)

                except ValueError:
                    return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

                # Obtener todas las cuentas ordenadas por código
                cuentas = PlanCuenta.objects.filter(
                    empresa__siglas__exact='PSM',
                    codigo__gte=cuenta_inicio,
                    codigo__lte=cuenta_fin
                ).order_by('codigo')

                if not cuentas.exists():
                    return JsonResponse({
                        'error': 'No se encontraron cuentas en el rango especificado'
                    }, status=404)

                data = []
                for cuenta in cuentas:
                    saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)

                    # Formatear el nombre con la indentación correcta
                    nombre_formateado = self.format_nombre_cuenta(cuenta.nombre, cuenta.codigo)

                    item = {
                        'id': cuenta.id,
                        'codigo': cuenta.codigo,
                        'nombre': nombre_formateado,
                        'tipo': self.get_account_type(cuenta),
                        **saldos
                    }
                    data.append(item)

                # Ordenar los datos por código
                data = sorted(data, key=lambda x: x['codigo'])

            elif action == 'get_cuentas':
                # Obtener todas las cuentas principales (nivel 1)
                cuentas_principales = PlanCuenta.objects.filter(
                    empresa__siglas__exact='PSM',
                    codigo__regex=r'^\d+$'  # Cuentas de nivel 1 (solo dígitos)
                ).order_by('codigo')

                data = [{'codigo': cuenta.codigo, 'nombre': cuenta.nombre}
                        for cuenta in cuentas_principales]
                return JsonResponse(data, safe=False)

            else:
                return JsonResponse({'error': 'Acción no reconocida'}, status=400)

        except Exception as e:
            import traceback
            logger.error(f"Error en post: {str(e)}\n{traceback.format_exc()}")
            return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Balance de Comprobación'
        context['title'] = 'Balance de Comprobación'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


# CODIGO SI VALE SOLO EL DE ARRIBA ESTA CON CAMBIO PARA MEJORAR LA SEPARACION POR FECHA DE CORTE
# class listarBalancePlanView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
#         """
#         Calcula los saldos de una cuenta para un período específico
#         """
#         # Saldo anterior: todas las transacciones hasta fecha_inicio (exclusive)
#         saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__lt=fecha_inicio
#         ).aggregate(
#             debe=Coalesce(Sum('debe'), Decimal('0')),
#             haber=Coalesce(Sum('haber'), Decimal('0'))
#         )
#
#         # Saldo del mes: transacciones entre fecha_inicio y fecha_fin (inclusive)
#         saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__gte=fecha_inicio,
#             encabezadocuentaplan__fecha__lte=fecha_fin
#         ).aggregate(
#             debe=Coalesce(Sum('debe'), Decimal('0')),
#             haber=Coalesce(Sum('haber'), Decimal('0'))
#         )
#
#         # Para cuentas padre, incluir saldos de subcuentas
#         subcuentas = PlanCuenta.objects.filter(parentId=cuenta)
#         if subcuentas.exists():
#             for subcuenta in subcuentas:
#                 sub_saldos = self.get_saldos_cuenta(subcuenta, fecha_inicio, fecha_fin)
#                 saldo_anterior['debe'] += Decimal(str(sub_saldos['debe_ant']))
#                 saldo_anterior['haber'] += Decimal(str(sub_saldos['haber_ant']))
#                 saldo_mes['debe'] += Decimal(str(sub_saldos['debe_mes']))
#                 saldo_mes['haber'] += Decimal(str(sub_saldos['haber_mes']))
#
#         debe_actual = saldo_anterior['debe'] + saldo_mes['debe']
#         haber_actual = saldo_anterior['haber'] + saldo_mes['haber']
#
#         return {
#             'debe_ant': float(saldo_anterior['debe']),
#             'haber_ant': float(saldo_anterior['haber']),
#             'debe_mes': float(saldo_mes['debe']),
#             'haber_mes': float(saldo_mes['haber']),
#             'debe_act': float(debe_actual),
#             'haber_act': float(haber_actual)
#         }
#
#     def get_account_type(self, cuenta):
#         """Determina el tipo de cuenta basado en su código y naturaleza"""
#         if len(cuenta.codigo) == 1:  # Cuenta mayor
#             return 'G'
#         elif len(cuenta.codigo) >= 9:  # Si tiene 9 o más caracteres es cuenta de detalle
#             return 'D'
#         else:
#             return 'G'
#
#     def get_nivel(self, codigo):
#         """Determina el nivel de la cuenta basado en su código"""
#         return len(codigo.split('.')[0])
#
#     def get_parent_code(self, codigo):
#         """Obtiene el código del padre de una cuenta"""
#         parts = codigo.split('.')
#         if len(parts[0]) > 1:
#             return parts[0][:-2] if len(parts[0]) > 2 else parts[0][:-1]
#         return None
#
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         try:
#             action = request.POST.get('action')
#             if action == 'searchdata_psm':
#                 fecha_inicio = request.POST.get('fecha_inicio')
#                 fecha_fin = request.POST.get('fecha_fin')
#                 cuenta_inicio = request.POST.get('cuenta_inicio', '1')
#                 cuenta_fin = request.POST.get('cuenta_fin', '9')
#
#                 if not fecha_inicio or not fecha_fin:
#                     return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)
#
#                 try:
#                     # Convertir las fechas asegurando que estén en el formato correcto
#                     fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
#                     fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
#
#                     # Validar que fecha_fin no sea menor que fecha_inicio
#                     if fecha_fin < fecha_inicio:
#                         return JsonResponse({
#                             'error': 'La fecha final no puede ser menor que la fecha inicial'
#                         }, status=400)
#
#                 except ValueError:
#                     return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
#
#                 # Obtener todas las cuentas ordenadas por código
#                 cuentas = PlanCuenta.objects.filter(
#                     empresa__siglas__exact='PSM',
#                     codigo__gte=cuenta_inicio,
#                     codigo__lte=cuenta_fin
#                 ).order_by('codigo')
#
#                 if not cuentas.exists():
#                     return JsonResponse({
#                         'error': 'No se encontraron cuentas en el rango especificado'
#                     }, status=404)
#
#                 data = []
#                 account_dict = {}
#
#                 # Primera pasada: calcular saldos para todas las cuentas
#                 for cuenta in cuentas:
#                     saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)
#                     nivel = self.get_nivel(cuenta.codigo)
#                     tipo_cuenta = self.get_account_type(cuenta)
#
#                     item = {
#                         'id': cuenta.id,
#                         'codigo': cuenta.codigo,
#                         'nombre': cuenta.nombre,
#                         'nivel': nivel,
#                         'tipo': tipo_cuenta,
#                         **saldos
#                     }
#                     account_dict[cuenta.codigo] = item
#
#                 # Convertir el diccionario a una lista ordenada
#                 data = sorted(account_dict.values(), key=lambda x: x['codigo'])
#
#             elif action == 'get_cuentas':
#                 # Obtener todas las cuentas principales (nivel 1)
#                 cuentas_principales = PlanCuenta.objects.filter(
#                     empresa__siglas__exact='PSM',
#                     codigo__regex=r'^\d+$'  # Cuentas de nivel 1 (solo dígitos)
#                 ).order_by('codigo')
#
#                 cuentas_data = [{'codigo': cuenta.codigo, 'nombre': cuenta.nombre}
#                                 for cuenta in cuentas_principales]
#                 return JsonResponse(cuentas_data, safe=False)
#
#             else:
#                 return JsonResponse({'error': 'Acción no reconocida'}, status=400)
#
#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)
#
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context


# class listarBalancePlanView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
#         saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__lt=fecha_inicio
#         ).aggregate(
#             debe=Sum('debe', default=0),
#             haber=Sum('haber', default=0)
#         )
#
#         saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
#             cuenta=cuenta,
#             encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin]
#         ).aggregate(
#             debe=Sum('debe', default=0),
#             haber=Sum('haber', default=0)
#         )
#
#         saldo_actual_debe = saldo_anterior['debe'] + saldo_mes['debe']
#         saldo_actual_haber = saldo_anterior['haber'] + saldo_mes['haber']
#
#         return {
#             'saldo_anterior_debe': float(saldo_anterior['debe']),
#             'saldo_anterior_haber': float(saldo_anterior['haber']),
#             'debe': float(saldo_mes['debe']),
#             'haber': float(saldo_mes['haber']),
#             'saldo_actual_debe': float(saldo_actual_debe),
#             'saldo_actual_haber': float(saldo_actual_haber)
#         }
#
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             action = request.POST.get('action')
#             if action == 'searchdata_psm':
#                 fecha_inicio = request.POST.get('fecha_inicio')
#                 fecha_fin = request.POST.get('fecha_fin')
#
#                 if not fecha_inicio or not fecha_fin:
#                     return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)
#
#                 try:
#                     fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
#                     fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
#                 except ValueError:
#                     return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
#
#                 cuentas = PlanCuenta.objects.filter(empresa__siglas__exact='PSM').order_by('codigo')
#
#                 data = []
#                 for cuenta in cuentas:
#                     saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)
#
#                     if (saldos['saldo_anterior_debe'] != 0 or
#                             saldos['saldo_anterior_haber'] != 0 or
#                             saldos['debe'] != 0 or
#                             saldos['haber'] != 0):
#                         item = {
#                             'cuenta': {
#                                 'codigo': cuenta.codigo,
#                                 'nombre': cuenta.nombre
#                             },
#                             **saldos
#                         }
#                         data.append(item)
#
#             elif action == 'searchdataplan':
#                 data = []
#                 for cuenta in PlanCuenta.objects.filter(empresa__siglas__exact='PSM').order_by('codigo'):
#                     data.append({
#                         'codigo': cuenta.codigo,
#                         'nombre': cuenta.nombre,
#                         'text': f"{cuenta.codigo} - {cuenta.nombre}"
#                     })
#
#             else:
#                 data = {'error': 'Acción no reconocida'}
#                 return JsonResponse(data, status=400)
#
#         except Exception as e:
#             data = {'error': str(e)}
#             return JsonResponse(data, status=500)
#
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context


# class listarBalancePlanView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_psm.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
#         try:
#             saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
#                 cuenta=cuenta,
#                 encabezadocuentaplan__fecha__lt=fecha_inicio
#             ).aggregate(
#                 debe=Sum('debe', default=0),
#                 haber=Sum('haber', default=0)
#             )
#
#             saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
#                 cuenta=cuenta,
#                 encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin]
#             ).aggregate(
#                 debe=Sum('debe', default=0),
#                 haber=Sum('haber', default=0)
#             )
#
#             return {
#                 'saldo_anterior_debe': float(saldo_anterior['debe'] or 0),
#                 'saldo_anterior_haber': float(saldo_anterior['haber'] or 0),
#                 'debe': float(saldo_mes['debe'] or 0),
#                 'haber': float(saldo_mes['haber'] or 0),
#                 'saldo_actual_debe': float((saldo_anterior['debe'] or 0) + (saldo_mes['debe'] or 0)),
#                 'saldo_actual_haber': float((saldo_anterior['haber'] or 0) + (saldo_mes['haber'] or 0))
#             }
#         except Exception as e:
#             print(f"Error calculando saldos para cuenta {cuenta.codigo}: {str(e)}")
#             return {
#                 'saldo_anterior_debe': 0,
#                 'saldo_anterior_haber': 0,
#                 'debe': 0,
#                 'haber': 0,
#                 'saldo_actual_debe': 0,
#                 'saldo_actual_haber': 0
#             }
#
#     def process_account_hierarchy(self, cuenta, fecha_inicio, fecha_fin, processed_accounts=None):
#         if processed_accounts is None:
#             processed_accounts = set()
#
#         if cuenta.codigo in processed_accounts:
#             return None
#
#         processed_accounts.add(cuenta.codigo)
#
#         saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)
#
#         # Solo incluir la cuenta si tiene movimientos o es una cuenta total
#         if (saldos['saldo_anterior_debe'] != 0 or
#                 saldos['saldo_anterior_haber'] != 0 or
#                 saldos['debe'] != 0 or
#                 saldos['haber'] != 0 or
#                 cuenta.band_total):
#
#             account_data = {
#                 'cuenta': {
#                     'codigo': cuenta.codigo,
#                     'nombre': cuenta.nombre,
#                     'nivel': cuenta.nivel
#                 },
#                 'is_total': cuenta.band_total,
#                 **saldos
#             }
#
#             # Procesar cuentas hijas
#             children = PlanCuenta.objects.filter(parentId=cuenta).order_by('codigo')
#             if children.exists():
#                 account_data['children'] = []
#                 for child in children:
#                     child_data = self.process_account_hierarchy(child, fecha_inicio, fecha_fin, processed_accounts)
#                     if child_data:
#                         account_data['children'].append(child_data)
#
#             return account_data
#         return None
#
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             action = request.POST.get('action')
#             if action == 'searchdata_psm':
#                 fecha_inicio = request.POST.get('fecha_inicio')
#                 fecha_fin = request.POST.get('fecha_fin')
#
#                 if not fecha_inicio or not fecha_fin:
#                     return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)
#
#                 try:
#                     self.fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
#                     self.fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
#                 except ValueError:
#                     return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
#
#                 # Obtener cuentas raíz (nivel 1)
#                 root_accounts = PlanCuenta.objects.filter(
#                     empresa__siglas__exact='PSM',
#                     nivel=1
#                 ).order_by('codigo')
#
#                 data = []
#                 processed_accounts = set()
#
#                 for root_account in root_accounts:
#                     account_data = self.process_account_hierarchy(
#                         root_account,
#                         self.fecha_inicio,
#                         self.fecha_fin,
#                         processed_accounts
#                     )
#                     if account_data:
#                         data.append(account_data)
#
#             elif action == 'searchdataplan':
#                 data = []
#                 for cuenta in PlanCuenta.objects.filter(
#                         empresa__siglas__exact='PSM'
#                 ).order_by('codigo'):
#                     data.append({
#                         'codigo': cuenta.codigo,
#                         'nombre': cuenta.nombre,
#                         'text': f"{cuenta.codigo} - {cuenta.nombre}"
#                     })
#             else:
#                 return JsonResponse({'error': 'Acción no reconocida'}, status=400)
#
#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)
#
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context

class listarBalancePlanBIOView(ListView):
    model = DetalleCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_bio.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_saldos_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        """
        Calcula los saldos de una cuenta para un período específico
        """
        # Saldo anterior: todas las transacciones hasta fecha_inicio (exclusive)
        saldo_anterior = DetalleCuentasPlanCuenta.objects.filter(
            cuenta=cuenta,
            encabezadocuentaplan__fecha__lt=fecha_inicio
        ).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        # Saldo del mes: transacciones entre fecha_inicio y fecha_fin (inclusive)
        saldo_mes = DetalleCuentasPlanCuenta.objects.filter(
            cuenta=cuenta,
            encabezadocuentaplan__fecha__gte=fecha_inicio,
            encabezadocuentaplan__fecha__lte=fecha_fin
        ).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        # Para cuentas padre, incluir saldos de subcuentas
        subcuentas = PlanCuenta.objects.filter(parentId=cuenta)
        if subcuentas.exists():
            for subcuenta in subcuentas:
                sub_saldos = self.get_saldos_cuenta(subcuenta, fecha_inicio, fecha_fin)
                saldo_anterior['debe'] += Decimal(str(sub_saldos['debe_ant']))
                saldo_anterior['haber'] += Decimal(str(sub_saldos['haber_ant']))
                saldo_mes['debe'] += Decimal(str(sub_saldos['debe_mes']))
                saldo_mes['haber'] += Decimal(str(sub_saldos['haber_mes']))

        debe_actual = saldo_anterior['debe'] + saldo_mes['debe']
        haber_actual = saldo_anterior['haber'] + saldo_mes['haber']

        return {
            'debe_ant': float(saldo_anterior['debe']),
            'haber_ant': float(saldo_anterior['haber']),
            'debe_mes': float(saldo_mes['debe']),
            'haber_mes': float(saldo_mes['haber']),
            'debe_act': float(debe_actual),
            'haber_act': float(haber_actual)
        }

    def get_account_type(self, cuenta):
        """Determina el tipo de cuenta basado en su código y naturaleza"""
        if len(cuenta.codigo) == 1:  # Cuenta mayor
            return 'G'
        elif len(cuenta.codigo) >= 9:  # Si tiene 9 o más caracteres es cuenta de detalle
            return 'D'
        else:
            return 'G'

    def get_nivel(self, codigo):
        """Determina el nivel de la cuenta basado en su código"""
        return len(codigo.split('.')[0])

    def get_parent_code(self, codigo):
        """Obtiene el código del padre de una cuenta"""
        parts = codigo.split('.')
        if len(parts[0]) > 1:
            return parts[0][:-2] if len(parts[0]) > 2 else parts[0][:-1]
        return None

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            action = request.POST.get('action')
            if action == 'searchdata_bio':
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin = request.POST.get('fecha_fin')
                cuenta_inicio = request.POST.get('cuenta_inicio', '1')
                cuenta_fin = request.POST.get('cuenta_fin', '9')

                if not fecha_inicio or not fecha_fin:
                    return JsonResponse({'error': 'Las fechas de inicio y fin son requeridas'}, status=400)

                try:
                    # Convertir las fechas asegurando que estén en el formato correcto
                    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()

                    # Validar que fecha_fin no sea menor que fecha_inicio
                    if fecha_fin < fecha_inicio:
                        return JsonResponse({
                            'error': 'La fecha final no puede ser menor que la fecha inicial'
                        }, status=400)

                except ValueError:
                    return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

                # Obtener todas las cuentas ordenadas por código
                cuentas = PlanCuenta.objects.filter(
                    empresa__siglas__exact='BIO',
                    codigo__gte=cuenta_inicio,
                    codigo__lte=cuenta_fin
                ).order_by('codigo')

                if not cuentas.exists():
                    return JsonResponse({
                        'error': 'No se encontraron cuentas en el rango especificado'
                    }, status=404)

                data = []
                account_dict = {}

                # Primera pasada: calcular saldos para todas las cuentas
                for cuenta in cuentas:
                    saldos = self.get_saldos_cuenta(cuenta, fecha_inicio, fecha_fin)
                    nivel = self.get_nivel(cuenta.codigo)
                    tipo_cuenta = self.get_account_type(cuenta)

                    item = {
                        'id': cuenta.id,
                        'codigo': cuenta.codigo,
                        'nombre': cuenta.nombre,
                        'nivel': nivel,
                        'tipo': tipo_cuenta,
                        **saldos
                    }
                    account_dict[cuenta.codigo] = item

                # Convertir el diccionario a una lista ordenada
                data = sorted(account_dict.values(), key=lambda x: x['codigo'])

            elif action == 'get_cuentas':
                # Obtener todas las cuentas principales (nivel 1)
                cuentas_principales = PlanCuenta.objects.filter(
                    empresa__siglas__exact='BIO',
                    codigo__regex=r'^\d+$'  # Cuentas de nivel 1 (solo dígitos)
                ).order_by('codigo')

                cuentas_data = [{'codigo': cuenta.codigo, 'nombre': cuenta.nombre} for cuenta in cuentas_principales]
                return JsonResponse(cuentas_data, safe=False)

            else:
                return JsonResponse({'error': 'Acción no reconocida'}, status=400)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'error': str(e)}, status=500)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Balance de Comprobación'
        context['title'] = 'Balance de Comprobación'
        context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
        return context


# class listarBalancePlanBIOView(ListView):
#     model = DetalleCuentasPlanCuenta
#     template_name = 'app_contabilidad_planCuentas/balance_Plan/balance_mayorizacion_bio.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         print('request.POST')
#         print(request.POST)
#         data = {}
#         try:
#             action = request.POST['action']
#             if action == 'searchdata_bio':
#                 data = []
#                 detallecuenta = DetalleCuentasPlanCuenta.objects.filter(cuenta__empresa__siglas__exact='BIO').order_by('cuenta__codigo')
#                 for i in detallecuenta:
#                     data.append(i.toJSON())
#
#             elif action == 'searchdataplan':
#                 print('llego a search data LIBRO DETALLE plan')
#                 data = []
#                 for i in PlanCuenta.objects.all().order_by('codigo'):
#                     item = i.toJSON()
#                     item['codigo'] = i.codigo
#                     item['text'] = i.get_name()
#                     data.append(item)
#
#             elif action == 'searchdatahierarchy':
#                 print('Generando jerarquía del plan de cuentas')
#
#                 # Construir jerarquía
#                 detallecuenta = DetalleCuentasPlanCuenta.objects.select_related('cuenta').all().order_by('cuenta__codigo')
#                 hierarchy = defaultdict(lambda: {"children": [], "cuenta": None})
#
#                 for item in detallecuenta:
#                     cuenta = item.cuenta
#                     codigo = cuenta.codigo
#                     parent_codigo = ".".join(codigo.split(".")[:-1]) if "." in codigo else None
#
#                     # Crear nodo
#                     node = {
#                         "codigo": cuenta.codigo,
#                         "nombre": cuenta.nombre,
#                         "debe": item.debe,
#                         "haber": item.haber,
#                         "saldo": item.debe - item.haber,  # Calcula el saldo
#                     }
#                     hierarchy[codigo]["cuenta"] = node
#
#                     # Agregar nodo al padre
#                     if parent_codigo:
#                         hierarchy[parent_codigo]["children"].append(hierarchy[codigo])
#
#                 # Filtrar nodos raíz
#                 data = [node for key, node in hierarchy.items() if "." not in key]
#
#             elif action == 'search_report':
#                 print('llego a search data RANGUE LIBRO DETALLE plan')
#                 data = []
#                 start_date = request.POST.get('start_date', '')
#                 end_date = request.POST.get('end_date', '')
#                 desde_rang = request.POST.get('desde_rang', '')
#                 hasta_rang = request.POST.get('hasta_rang', '')
#                 search = DetalleCuentasPlanCuenta.objects.all()
#                 if len(start_date) and len(end_date):
#                     search = search.filter(
#                         encabezadocuentaplan__fecha__range=[start_date, end_date],
#                         cuenta__codigo__range=[desde_rang, hasta_rang],
#                     )
#                 for i in search:
#                     data.append(i.toJSON())
#
#             else:
#                 data['error'] = 'Ha ocurrido un error'
#         except Exception as e:
#             data['error'] = str(e)
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Balance de Comprobación'
#         context['title'] = 'Balance de Comprobación'
#         context['list_url'] = reverse_lazy('app_planCuentas:listar_transaccionPlan')
#         return context


class DiarioGeneralAcumuladoPSMView(ListView):
    model = EncabezadoCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/diario_general/diario_acumuladoPSM.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                empresa = request.POST['empresa']
                fecha_inicio = request.POST.get('fecha_inicio', '2024-12-01')
                fecha_fin = request.POST.get('fecha_fin', '2025-01-31')

                asientos = DetalleCuentasPlanCuenta.objects.select_related(
                    'cuenta',
                    'encabezadocuentaplan'
                ).filter(
                    encabezadocuentaplan__empresa__siglas__exact=empresa,
                    encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin]
                ).order_by(
                    'encabezadocuentaplan__fecha',
                    'encabezadocuentaplan__id',
                    'orden'
                )

                data = self.process_asientos(asientos)
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def process_asientos(self, asientos):
        total_general_debe = 0
        total_general_haber = 0
        fecha_actual = None
        asientos_agrupados = []
        subtotal_debe = 0
        subtotal_haber = 0

        for asiento in asientos:
            if fecha_actual != asiento.encabezadocuentaplan.fecha:
                if fecha_actual is not None:
                    asientos_agrupados.append({
                        'tipo': 'total',
                        'fecha': fecha_actual.strftime('%Y-%m-%d'),
                        'codigo': '',
                        'nombre': '',
                        'doc': '',
                        'descripcion': 'TOTAL FECHA ................',
                        'debe': float(subtotal_debe),
                        'haber': float(subtotal_haber)
                    })
                fecha_actual = asiento.encabezadocuentaplan.fecha
                subtotal_debe = 0
                subtotal_haber = 0

            asientos_agrupados.append({
                'tipo': 'asiento',
                'fecha': asiento.encabezadocuentaplan.fecha.strftime('%Y-%m-%d'),
                'codigo': asiento.cuenta.codigo,
                'nombre': asiento.cuenta.nombre,
                'doc': asiento.encabezadocuentaplan.comprobante,
                'descripcion': asiento.detalle or asiento.encabezadocuentaplan.descripcion,
                'debe': float(asiento.debe),
                'haber': float(asiento.haber)
            })

            subtotal_debe += asiento.debe
            subtotal_haber += asiento.haber
            total_general_debe += asiento.debe
            total_general_haber += asiento.haber

        if fecha_actual is not None:
            asientos_agrupados.append({
                'tipo': 'total',
                'fecha': fecha_actual.strftime('%Y-%m-%d'),
                'codigo': '',
                'nombre': '',
                'doc': '',
                'descripcion': 'TOTAL FECHA ................',
                'debe': float(subtotal_debe),
                'haber': float(subtotal_haber)
            })

        return {
            'asientos': asientos_agrupados,
            'total_general': {
                'tipo': 'grand_total',
                'fecha': '',
                'codigo': '',
                'nombre': '',
                'doc': '',
                'descripcion': 'TOTAL DEL REPORTE....',
                'debe': float(total_general_debe),
                'haber': float(total_general_haber)
            }
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Diario General Acumulado Empresa PSM'
        context['list_url'] = reverse_lazy('diario_general:diario_acumulado')
        # context['empresa_id'] = self.kwargs['empresa_id']
        # context['empresa_nombre'] = Empresa.objects.get(id=self.kwargs['empresa_id']).nombre
        return context


class DiarioGeneralAcumuladoBIOView(ListView):
    model = EncabezadoCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/diario_general/diario_acumuladoBIO.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'searchdata':
                empresa = request.POST['empresa']
                print('empresa')
                print(empresa)
                fecha_inicio = request.POST.get('fecha_inicio', '2024-12-01')
                fecha_fin = request.POST.get('fecha_fin', '2025-01-31')

                asientos = DetalleCuentasPlanCuenta.objects.select_related(
                    'cuenta',
                    'encabezadocuentaplan'
                ).filter(
                    encabezadocuentaplan__empresa__siglas__exact=empresa,
                    encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin]
                ).order_by(
                    'encabezadocuentaplan__fecha',
                    'encabezadocuentaplan__id',
                    'orden'
                )

                data = self.process_asientos(asientos)
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def process_asientos(self, asientos):
        total_general_debe = 0
        total_general_haber = 0
        fecha_actual = None
        asientos_agrupados = []
        subtotal_debe = 0
        subtotal_haber = 0

        for asiento in asientos:
            if fecha_actual != asiento.encabezadocuentaplan.fecha:
                if fecha_actual is not None:
                    asientos_agrupados.append({
                        'tipo': 'total',
                        'fecha': fecha_actual.strftime('%Y-%m-%d'),
                        'codigo': '',
                        'nombre': '',
                        'doc': '',
                        'descripcion': 'TOTAL FECHA ................',
                        'debe': float(subtotal_debe),
                        'haber': float(subtotal_haber)
                    })
                fecha_actual = asiento.encabezadocuentaplan.fecha
                subtotal_debe = 0
                subtotal_haber = 0

            asientos_agrupados.append({
                'tipo': 'asiento',
                'fecha': asiento.encabezadocuentaplan.fecha.strftime('%Y-%m-%d'),
                'codigo': asiento.cuenta.codigo,
                'nombre': asiento.cuenta.nombre,
                'doc': asiento.encabezadocuentaplan.comprobante,
                'descripcion': asiento.detalle or asiento.encabezadocuentaplan.descripcion,
                'debe': float(asiento.debe),
                'haber': float(asiento.haber)
            })

            subtotal_debe += asiento.debe
            subtotal_haber += asiento.haber
            total_general_debe += asiento.debe
            total_general_haber += asiento.haber

        if fecha_actual is not None:
            asientos_agrupados.append({
                'tipo': 'total',
                'fecha': fecha_actual.strftime('%Y-%m-%d'),
                'codigo': '',
                'nombre': '',
                'doc': '',
                'descripcion': 'TOTAL FECHA ................',
                'debe': float(subtotal_debe),
                'haber': float(subtotal_haber)
            })

        return {
            'asientos': asientos_agrupados,
            'total_general': {
                'tipo': 'grand_total',
                'fecha': '',
                'codigo': '',
                'nombre': '',
                'doc': '',
                'descripcion': 'TOTAL DEL REPORTE....',
                'debe': float(total_general_debe),
                'haber': float(total_general_haber)
            }
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Diario General Acumulado Empresa BIO'
        context['list_url'] = reverse_lazy('diario_general:diario_acumulado')
        # context['empresa_id'] = self.kwargs['empresa_id']
        # context['empresa_nombre'] = Empresa.objects.get(id=self.kwargs['empresa_id']).nombre
        return context


# =========================
# LIBRO MAYOR - DETALLE POR CUENTA
# =========================
class LibroMayorDetalleView(ListView):
    model = DetalleCuentasPlanCuenta
    template_name = 'app_contabilidad_planCuentas/libro_mayor/libro_mayor_detalle.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'search_mayor':
                data = []
                cuenta_id = request.POST.get('cuenta_id')
                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin = request.POST.get('fecha_fin')
                empresa_siglas = request.POST.get('empresa', 'PSM')

                queryset = DetalleCuentasPlanCuenta.objects.filter(
                    encabezadocuentaplan__empresa__siglas=empresa_siglas,
                    encabezadocuentaplan__reg_control='RT'
                ).select_related('cuenta', 'encabezadocuentaplan')

                if cuenta_id:
                    queryset = queryset.filter(cuenta_id=cuenta_id)
                if fecha_inicio:
                    queryset = queryset.filter(encabezadocuentaplan__fecha__gte=fecha_inicio)
                if fecha_fin:
                    queryset = queryset.filter(encabezadocuentaplan__fecha__lte=fecha_fin)

                queryset = queryset.order_by('encabezadocuentaplan__fecha', 'encabezadocuentaplan__codigo')

                saldo_acumulado = Decimal('0')
                for det in queryset:
                    # Calcular saldo segun naturaleza de la cuenta
                    if det.cuenta.band_deudor:  # Cuenta deudora
                        saldo_acumulado += det.debe - det.haber
                    else:  # Cuenta acreedora
                        saldo_acumulado += det.haber - det.debe

                    item = det.toJSON()
                    item['saldo_acumulado'] = format(saldo_acumulado, '.2f')
                    item['fecha_transaccion'] = det.encabezadocuentaplan.fecha.strftime(
                        '%Y-%m-%d') if det.encabezadocuentaplan.fecha else ''
                    item['codigo_transaccion'] = det.encabezadocuentaplan.codigo
                    item['descripcion_transaccion'] = det.encabezadocuentaplan.descripcion
                    item['comprobante'] = det.encabezadocuentaplan.comprobante
                    data.append(item)

            elif action == 'get_cuentas':
                data = []
                empresa_siglas = request.POST.get('empresa', 'PSM')
                cuentas = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')
                for c in cuentas:
                    data.append({
                        'id': c.id,
                        'codigo': c.codigo,
                        'nombre': c.nombre,
                        'full_name': f'{c.codigo} - {c.nombre}'
                    })

            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Libro Mayor'
        context['title'] = 'Libro Mayor - Detalle por Cuenta'
        context['cuentas'] = PlanCuenta.objects.filter(tipo_cuenta='DETALLE').order_by('codigo')
        return context


# =========================
# ESTADO DE RESULTADOS
# =========================
# class EstadoResultadosView(TemplateView):
#     template_name = 'app_contabilidad_planCuentas/reportes/estado_resultados.html'
#
#     @method_decorator(csrf_exempt)
#     @method_decorator(login_required)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             action = request.POST.get('action')
#
#             if action == 'generar_estado_resultados':
#                 fecha_inicio = request.POST.get('fecha_inicio')
#                 fecha_fin = request.POST.get('fecha_fin')
#                 empresa_siglas = request.POST.get('empresa', 'PSM')
#
#                 # Obtener cuentas de INGRESOS (codigo empieza con 4)
#                 ingresos = PlanCuenta.objects.filter(
#                     empresa__siglas=empresa_siglas,
#                     codigo__startswith='4',
#                     tipo_cuenta='DETALLE'
#                 ).order_by('codigo')
#
#                 # Obtener cuentas de GASTOS/COSTOS (codigo empieza con 5 o 6)
#                 gastos = PlanCuenta.objects.filter(
#                     empresa__siglas=empresa_siglas,
#                     tipo_cuenta='DETALLE'
#                 ).filter(
#                     Q(codigo__startswith='5') | Q(codigo__startswith='6')
#                 ).order_by('codigo')
#
#                 data_ingresos = []
#                 total_ingresos = Decimal('0')
#
#                 for cuenta in ingresos:
#                     saldos = self.calcular_saldo_cuenta(cuenta, fecha_inicio, fecha_fin)
#                     # Ingresos son acreedores: saldo = haber - debe
#                     saldo = saldos['haber'] - saldos['debe']
#                     if saldo != 0:
#                         data_ingresos.append({
#                             'codigo': cuenta.codigo,
#                             'nombre': cuenta.nombre,
#                             'saldo': format(saldo, '.2f')
#                         })
#                         total_ingresos += saldo
#
#                 data_gastos = []
#                 total_gastos = Decimal('0')
#
#                 for cuenta in gastos:
#                     saldos = self.calcular_saldo_cuenta(cuenta, fecha_inicio, fecha_fin)
#                     # Gastos son deudores: saldo = debe - haber
#                     saldo = saldos['debe'] - saldos['haber']
#                     if saldo != 0:
#                         data_gastos.append({
#                             'codigo': cuenta.codigo,
#                             'nombre': cuenta.nombre,
#                             'saldo': format(saldo, '.2f')
#                         })
#                         total_gastos += saldo
#
#                 utilidad_neta = total_ingresos - total_gastos
#
#                 data = {
#                     'ingresos': data_ingresos,
#                     'gastos': data_gastos,
#                     'total_ingresos': format(total_ingresos, '.2f'),
#                     'total_gastos': format(total_gastos, '.2f'),
#                     'utilidad_neta': format(utilidad_neta, '.2f'),
#                     'es_utilidad': utilidad_neta >= 0
#                 }
#             else:
#                 data['error'] = 'Accion no valida'
#
#         except Exception as e:
#             data['error'] = str(e)
#         return JsonResponse(data, safe=False)
#
#     def calcular_saldo_cuenta(self, cuenta, fecha_inicio, fecha_fin):
#         filtros = {
#             'cuenta': cuenta,
#             'encabezadocuentaplan__reg_control': 'RT'
#         }
#         if fecha_inicio:
#             filtros['encabezadocuentaplan__fecha__gte'] = fecha_inicio
#         if fecha_fin:
#             filtros['encabezadocuentaplan__fecha__lte'] = fecha_fin
#
#         totales = DetalleCuentasPlanCuenta.objects.filter(**filtros).aggregate(
#             debe=Coalesce(Sum('debe'), Decimal('0')),
#             haber=Coalesce(Sum('haber'), Decimal('0'))
#         )
#         return totales
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['nombre'] = 'Estado de Resultados'
#         context['title'] = 'Estado de Resultados'
#         return context


# =========================
# BALANCE GENERAL
# =========================
class BalanceGeneralView(TemplateView):
    template_name = 'app_contabilidad_planCuentas/reportes/balance_general.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'generar_balance_general':
                fecha_corte = request.POST.get('fecha_corte')
                empresa_siglas = request.POST.get('empresa', 'PSM')

                # ACTIVOS (codigo empieza con 1)
                activos = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='1',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                # PASIVOS (codigo empieza con 2)
                pasivos = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='2',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                # PATRIMONIO (codigo empieza con 3)
                patrimonio = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='3',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                data_activos = []
                total_activos = Decimal('0')

                for cuenta in activos:
                    saldos = self.calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    # Activos son deudores: saldo = debe - haber
                    saldo = saldos['debe'] - saldos['haber']
                    if saldo != 0:
                        data_activos.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(saldo, '.2f')
                        })
                        total_activos += saldo

                data_pasivos = []
                total_pasivos = Decimal('0')

                for cuenta in pasivos:
                    saldos = self.calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    # Pasivos son acreedores: saldo = haber - debe
                    saldo = saldos['haber'] - saldos['debe']
                    if saldo != 0:
                        data_pasivos.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(saldo, '.2f')
                        })
                        total_pasivos += saldo

                data_patrimonio = []
                total_patrimonio = Decimal('0')

                for cuenta in patrimonio:
                    saldos = self.calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    # Patrimonio es acreedor: saldo = haber - debe
                    saldo = saldos['haber'] - saldos['debe']
                    if saldo != 0:
                        data_patrimonio.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(saldo, '.2f')
                        })
                        total_patrimonio += saldo

                total_pasivo_patrimonio = total_pasivos + total_patrimonio
                diferencia = total_activos - total_pasivo_patrimonio
                cuadrado = abs(diferencia) < Decimal('0.01')

                data = {
                    'activos': data_activos,
                    'pasivos': data_pasivos,
                    'patrimonio': data_patrimonio,
                    'total_activos': format(total_activos, '.2f'),
                    'total_pasivos': format(total_pasivos, '.2f'),
                    'total_patrimonio': format(total_patrimonio, '.2f'),
                    'total_pasivo_patrimonio': format(total_pasivo_patrimonio, '.2f'),
                    'cuadrado': cuadrado,
                    'diferencia': format(diferencia, '.2f')
                }
            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def calcular_saldo_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        filtros = {
            'cuenta': cuenta,
            'encabezadocuentaplan__reg_control': 'RT'
        }
        if fecha_inicio:
            filtros['encabezadocuentaplan__fecha__gte'] = fecha_inicio
        if fecha_fin:
            filtros['encabezadocuentaplan__fecha__lte'] = fecha_fin

        totales = DetalleCuentasPlanCuenta.objects.filter(**filtros).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )
        return totales

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Balance General'
        context['title'] = 'Balance General'
        return context


# =========================
# ESTADO DE RESULTADOS - VISTA BASE
# =========================
class EstadoResultadosView(TemplateView):
    template_name = 'app_contabilidad_planCuentas/reportes/estado_resultados.html'

    def get_empresa_default(self):
        return 'PSM'

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}

        try:
            action = request.POST.get('action')

            if action == 'generar_estado_resultados':

                fecha_inicio = request.POST.get('fecha_inicio')
                fecha_fin = request.POST.get('fecha_fin')
                empresa_siglas = request.POST.get('empresa') or self.get_empresa_default()

                # 🔥 QUERY REAL (UNA SOLA)
                detalles = DetalleCuentasPlanCuenta.objects.filter(
                    encabezadocuentaplan__fecha__range=[fecha_inicio, fecha_fin],
                    encabezadocuentaplan__empresa__siglas=empresa_siglas
                ).values(
                    'cuenta__codigo',
                    'cuenta__nombre'
                ).annotate(
                    total_debe=Coalesce(Sum('debe'), Decimal('0')),
                    total_haber=Coalesce(Sum('haber'), Decimal('0'))
                )

                ingresos = []
                gastos = []

                total_ingresos = Decimal('0')
                total_gastos = Decimal('0')

                # 🔥 PROCESAMIENTO
                for d in detalles:

                    codigo = d['cuenta__codigo']
                    nombre = d['cuenta__nombre']

                    debe = d['total_debe']
                    haber = d['total_haber']

                    print("DEBUG REAL:", codigo, debe, haber)

                    # ================= INGRESOS =================
                    if codigo.startswith('4'):
                        saldo = haber - debe

                        if saldo != 0:
                            ingresos.append({
                                'codigo': codigo,
                                'nombre': nombre,
                                'saldo': format(abs(saldo), '.2f')
                            })
                            total_ingresos += abs(saldo)

                    # ================= GASTOS =================
                    elif codigo.startswith('5') or codigo.startswith('6'):
                        saldo = debe - haber

                        if saldo != 0:
                            gastos.append({
                                'codigo': codigo,
                                'nombre': nombre,
                                'saldo': format(abs(saldo), '.2f')
                            })
                            total_gastos += abs(saldo)

                # ================= RESULTADO =================
                utilidad_neta = total_ingresos - total_gastos

                data = {
                    'ingresos': ingresos,
                    'gastos': gastos,
                    'total_ingresos': format(total_ingresos, '.2f'),
                    'total_gastos': format(total_gastos, '.2f'),
                    'utilidad_neta': format(utilidad_neta, '.2f'),
                    'es_utilidad': utilidad_neta >= 0
                }

            else:
                data['error'] = 'Acción no válida'

        except Exception as e:
            data['error'] = str(e)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Estado de Resultados'
        context['empresa_default'] = self.get_empresa_default()
        return context


# ================= BIO =================
class EstadoResultadosBIOView(EstadoResultadosView):
    def get_empresa_default(self):
        return 'BIO'


# ================= PSM =================
class EstadoResultadosPSMView(EstadoResultadosView):
    def get_empresa_default(self):
        return 'PSM'


# =========================
# BALANCE GENERAL
# =========================
class BalanceGeneralView(TemplateView):
    template_name = 'app_contabilidad_planCuentas/reportes/balance_general.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'generar_balance_general':
                fecha_corte = request.POST.get('fecha_corte')
                empresa_siglas = request.POST.get('empresa', 'PSM')

                # ========== ACTIVOS (codigo empieza con 1) ==========
                activos_corrientes = []
                activos_no_corrientes = []
                total_activos_corrientes = Decimal('0')
                total_activos_no_corrientes = Decimal('0')

                # Activos Corrientes (11)
                cuentas_ac = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='11',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                for cuenta in cuentas_ac:
                    saldos = self._calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    saldo = saldos['debe'] - saldos['haber']  # Activos son deudores
                    if saldo != 0:
                        activos_corrientes.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(abs(saldo), '.2f')
                        })
                        total_activos_corrientes += abs(saldo)

                # Activos No Corrientes (12, 13, 14, etc.)
                cuentas_anc = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='1',
                    tipo_cuenta='DETALLE'
                ).exclude(codigo__startswith='11').order_by('codigo')

                for cuenta in cuentas_anc:
                    saldos = self._calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    saldo = saldos['debe'] - saldos['haber']
                    if saldo != 0:
                        activos_no_corrientes.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(abs(saldo), '.2f')
                        })
                        total_activos_no_corrientes += abs(saldo)

                total_activos = total_activos_corrientes + total_activos_no_corrientes

                # ========== PASIVOS (codigo empieza con 2) ==========
                pasivos_corrientes = []
                pasivos_no_corrientes = []
                total_pasivos_corrientes = Decimal('0')
                total_pasivos_no_corrientes = Decimal('0')

                # Pasivos Corrientes (21)
                cuentas_pc = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='21',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                for cuenta in cuentas_pc:
                    saldos = self._calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    saldo = saldos['haber'] - saldos['debe']  # Pasivos son acreedores
                    if saldo != 0:
                        pasivos_corrientes.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(abs(saldo), '.2f')
                        })
                        total_pasivos_corrientes += abs(saldo)

                # Pasivos No Corrientes (22, 23, etc.)
                cuentas_pnc = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='2',
                    tipo_cuenta='DETALLE'
                ).exclude(codigo__startswith='21').order_by('codigo')

                for cuenta in cuentas_pnc:
                    saldos = self._calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    saldo = saldos['haber'] - saldos['debe']
                    if saldo != 0:
                        pasivos_no_corrientes.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(abs(saldo), '.2f')
                        })
                        total_pasivos_no_corrientes += abs(saldo)

                total_pasivos = total_pasivos_corrientes + total_pasivos_no_corrientes

                # ========== PATRIMONIO (codigo empieza con 3) ==========
                patrimonio_data = []
                total_patrimonio = Decimal('0')

                cuentas_patrimonio = PlanCuenta.objects.filter(
                    empresa__siglas=empresa_siglas,
                    codigo__startswith='3',
                    tipo_cuenta='DETALLE'
                ).order_by('codigo')

                for cuenta in cuentas_patrimonio:
                    saldos = self._calcular_saldo_cuenta(cuenta, None, fecha_corte)
                    saldo = saldos['haber'] - saldos['debe']  # Patrimonio es acreedor
                    if saldo != 0:
                        patrimonio_data.append({
                            'codigo': cuenta.codigo,
                            'nombre': cuenta.nombre,
                            'saldo': format(abs(saldo), '.2f')
                        })
                        total_patrimonio += abs(saldo)

                # VERIFICACION ECUACION CONTABLE
                total_pasivo_patrimonio = total_pasivos + total_patrimonio
                diferencia = total_activos - total_pasivo_patrimonio
                cuadrado = abs(diferencia) < Decimal('0.01')

                data = {
                    'activos_corrientes': activos_corrientes,
                    'activos_no_corrientes': activos_no_corrientes,
                    'pasivos_corrientes': pasivos_corrientes,
                    'pasivos_no_corrientes': pasivos_no_corrientes,
                    'patrimonio': patrimonio_data,
                    'total_activos_corrientes': format(total_activos_corrientes, '.2f'),
                    'total_activos_no_corrientes': format(total_activos_no_corrientes, '.2f'),
                    'total_activos': format(total_activos, '.2f'),
                    'total_pasivos_corrientes': format(total_pasivos_corrientes, '.2f'),
                    'total_pasivos_no_corrientes': format(total_pasivos_no_corrientes, '.2f'),
                    'total_pasivos': format(total_pasivos, '.2f'),
                    'total_patrimonio': format(total_patrimonio, '.2f'),
                    'total_pasivo_patrimonio': format(total_pasivo_patrimonio, '.2f'),
                    'cuadrado': cuadrado,
                    'diferencia': format(diferencia, '.2f')
                }
            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def _calcular_saldo_cuenta(self, cuenta, fecha_inicio, fecha_fin):
        filtros = {
            'cuenta': cuenta,
            'encabezadocuentaplan__reg_control': 'RT'
        }
        if fecha_inicio:
            filtros['encabezadocuentaplan__fecha__gte'] = fecha_inicio
        if fecha_fin:
            filtros['encabezadocuentaplan__fecha__lte'] = fecha_fin

        totales = DetalleCuentasPlanCuenta.objects.filter(**filtros).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )
        return totales

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Balance General'
        context['title'] = 'Balance General'
        return context


class BalanceGeneralBIOView(BalanceGeneralView):
    template_name = 'app_contabilidad_planCuentas/reportes/balance_general_bio.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Balance General - BIO'
        context['empresa_default'] = 'BIO'
        return context


# =========================
# CIERRE CONTABLE
# =========================
class CierreContableView(TemplateView):
    template_name = 'app_contabilidad_planCuentas/reportes/cierre_contable.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'preview_cierre':
                # Vista previa del cierre sin ejecutar
                fecha_cierre = request.POST.get('fecha_cierre')
                empresa_siglas = request.POST.get('empresa', 'PSM')
                anio = request.POST.get('anio')

                # Obtener saldos de cuentas de resultado (4, 5, 6)
                cuentas_ingresos = self._obtener_saldos_cuentas('4', empresa_siglas, anio)
                cuentas_gastos = self._obtener_saldos_cuentas('5', empresa_siglas, anio)
                cuentas_costos = self._obtener_saldos_cuentas('6', empresa_siglas, anio)

                total_ingresos = sum(c['saldo'] for c in cuentas_ingresos)
                total_gastos = sum(c['saldo'] for c in cuentas_gastos)
                total_costos = sum(c['saldo'] for c in cuentas_costos)

                utilidad_perdida = total_ingresos - total_gastos - total_costos

                data = {
                    'cuentas_ingresos': cuentas_ingresos,
                    'cuentas_gastos': cuentas_gastos,
                    'cuentas_costos': cuentas_costos,
                    'total_ingresos': format(total_ingresos, '.2f'),
                    'total_gastos': format(total_gastos, '.2f'),
                    'total_costos': format(total_costos, '.2f'),
                    'utilidad_perdida': format(utilidad_perdida, '.2f'),
                    'es_utilidad': utilidad_perdida >= 0
                }

            elif action == 'ejecutar_cierre':
                with transaction.atomic():
                    fecha_cierre = request.POST.get('fecha_cierre')
                    empresa_siglas = request.POST.get('empresa', 'PSM')
                    anio = request.POST.get('anio')

                    empresa = Empresa.objects.get(siglas=empresa_siglas)

                    # Obtener cuenta de Utilidad/Perdida del Ejercicio (normalmente 3.6.x)
                    cuenta_resultado = PlanCuenta.objects.filter(
                        empresa=empresa,
                        codigo__startswith='36',
                        tipo_cuenta='DETALLE'
                    ).first()

                    if not cuenta_resultado:
                        raise Exception('No se encontro la cuenta de Utilidad/Perdida del Ejercicio (36.x.x)')

                    # Calcular utilidad/perdida
                    cuentas_ingresos = self._obtener_saldos_cuentas('4', empresa_siglas, anio)
                    cuentas_gastos = self._obtener_saldos_cuentas('5', empresa_siglas, anio)
                    cuentas_costos = self._obtener_saldos_cuentas('6', empresa_siglas, anio)

                    total_ingresos = sum(Decimal(str(c['saldo'])) for c in cuentas_ingresos)
                    total_gastos = sum(Decimal(str(c['saldo'])) for c in cuentas_gastos)
                    total_costos = sum(Decimal(str(c['saldo'])) for c in cuentas_costos)

                    utilidad_perdida = total_ingresos - total_gastos - total_costos

                    # Crear asiento de cierre
                    ultimo_codigo = EncabezadoCuentasPlanCuenta.objects.filter(
                        empresa=empresa
                    ).order_by('-codigo').first()

                    nuevo_codigo = (ultimo_codigo.codigo + 1) if ultimo_codigo else 1

                    encabezado = EncabezadoCuentasPlanCuenta.objects.create(
                        codigo=nuevo_codigo,
                        tip_cuenta='CIERRE',
                        tip_transa='CIERRE CONTABLE',
                        fecha=fecha_cierre,
                        comprobante=f'CIERRE-{anio}',
                        descripcion=f'Asiento de cierre contable periodo {anio}',
                        empresa=empresa,
                        reg_control='RT'
                    )

                    # Cerrar cuentas de INGRESOS (debitar para saldar)
                    for cuenta_data in cuentas_ingresos:
                        if cuenta_data['saldo'] != 0:
                            cuenta = PlanCuenta.objects.get(id=cuenta_data['id'])
                            DetalleCuentasPlanCuenta.objects.create(
                                encabezadocuentaplan=encabezado,
                                cuenta=cuenta,
                                detalle=f'Cierre cuenta {cuenta.codigo}',
                                debe=Decimal(str(abs(cuenta_data['saldo']))),
                                haber=Decimal('0')
                            )

                    # Cerrar cuentas de GASTOS (acreditar para saldar)
                    for cuenta_data in cuentas_gastos:
                        if cuenta_data['saldo'] != 0:
                            cuenta = PlanCuenta.objects.get(id=cuenta_data['id'])
                            DetalleCuentasPlanCuenta.objects.create(
                                encabezadocuentaplan=encabezado,
                                cuenta=cuenta,
                                detalle=f'Cierre cuenta {cuenta.codigo}',
                                debe=Decimal('0'),
                                haber=Decimal(str(abs(cuenta_data['saldo'])))
                            )

                    # Cerrar cuentas de COSTOS (acreditar para saldar)
                    for cuenta_data in cuentas_costos:
                        if cuenta_data['saldo'] != 0:
                            cuenta = PlanCuenta.objects.get(id=cuenta_data['id'])
                            DetalleCuentasPlanCuenta.objects.create(
                                encabezadocuentaplan=encabezado,
                                cuenta=cuenta,
                                detalle=f'Cierre cuenta {cuenta.codigo}',
                                debe=Decimal('0'),
                                haber=Decimal(str(abs(cuenta_data['saldo'])))
                            )

                    # Registrar utilidad/perdida en cuenta de patrimonio
                    if utilidad_perdida >= 0:
                        # Utilidad: se acredita (aumenta patrimonio)
                        DetalleCuentasPlanCuenta.objects.create(
                            encabezadocuentaplan=encabezado,
                            cuenta=cuenta_resultado,
                            detalle=f'Utilidad del ejercicio {anio}',
                            debe=Decimal('0'),
                            haber=utilidad_perdida
                        )
                    else:
                        # Perdida: se debita (disminuye patrimonio)
                        DetalleCuentasPlanCuenta.objects.create(
                            encabezadocuentaplan=encabezado,
                            cuenta=cuenta_resultado,
                            detalle=f'Perdida del ejercicio {anio}',
                            debe=abs(utilidad_perdida),
                            haber=Decimal('0')
                        )

                    data = {
                        'success': True,
                        'mensaje': f'Cierre contable ejecutado correctamente. Asiento #{nuevo_codigo}',
                        'asiento_id': encabezado.pk
                    }

            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def _obtener_saldos_cuentas(self, prefijo_codigo, empresa_siglas, anio):
        cuentas = PlanCuenta.objects.filter(
            empresa__siglas=empresa_siglas,
            codigo__startswith=prefijo_codigo,
            tipo_cuenta='DETALLE'
        ).order_by('codigo')

        resultado = []
        for cuenta in cuentas:
            saldos = DetalleCuentasPlanCuenta.objects.filter(
                cuenta=cuenta,
                encabezadocuentaplan__reg_control='RT',
                encabezadocuentaplan__fecha__year=anio
            ).aggregate(
                debe=Coalesce(Sum('debe'), Decimal('0')),
                haber=Coalesce(Sum('haber'), Decimal('0'))
            )

            # Calcular saldo segun naturaleza
            if cuenta.band_deudor:  # Deudora (gastos, costos)
                saldo = float(saldos['debe'] - saldos['haber'])
            else:  # Acreedora (ingresos)
                saldo = float(saldos['haber'] - saldos['debe'])

            if saldo != 0:
                resultado.append({
                    'id': cuenta.id,
                    'codigo': cuenta.codigo,
                    'nombre': cuenta.nombre,
                    'saldo': saldo
                })

        return resultado

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Cierre Contable'
        context['title'] = 'Proceso de Cierre Contable'
        context['anios'] = list(range(datetime.now().year, 2020, -1))
        return context


# =========================
# RATIOS FINANCIEROS
# =========================
class RatiosFinancierosView(TemplateView):
    template_name = 'app_contabilidad_planCuentas/reportes/ratios_financieros.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'calcular_ratios':
                fecha_corte = request.POST.get('fecha_corte')
                empresa_siglas = request.POST.get('empresa', 'PSM')

                # Obtener totales de Balance General
                activo_corriente = self._sumar_cuentas('11', empresa_siglas, fecha_corte, 'deudor')
                activo_no_corriente = self._sumar_cuentas('12', empresa_siglas, fecha_corte, 'deudor') + \
                                      self._sumar_cuentas('13', empresa_siglas, fecha_corte, 'deudor') + \
                                      self._sumar_cuentas('14', empresa_siglas, fecha_corte, 'deudor')
                total_activo = activo_corriente + activo_no_corriente

                pasivo_corriente = self._sumar_cuentas('21', empresa_siglas, fecha_corte, 'acreedor')
                pasivo_no_corriente = self._sumar_cuentas('22', empresa_siglas, fecha_corte, 'acreedor') + \
                                      self._sumar_cuentas('23', empresa_siglas, fecha_corte, 'acreedor')
                total_pasivo = pasivo_corriente + pasivo_no_corriente

                total_patrimonio = self._sumar_cuentas('3', empresa_siglas, fecha_corte, 'acreedor')

                # Obtener cuentas especificas
                caja_bancos = self._sumar_cuentas('1101', empresa_siglas, fecha_corte, 'deudor') + \
                              self._sumar_cuentas('1102', empresa_siglas, fecha_corte, 'deudor')
                inventarios = self._sumar_cuentas('1103', empresa_siglas, fecha_corte, 'deudor') + \
                              self._sumar_cuentas('1104', empresa_siglas, fecha_corte, 'deudor')
                cuentas_cobrar = self._sumar_cuentas('1105', empresa_siglas, fecha_corte, 'deudor')

                # Estado de Resultados (ultimo anio)
                anio = datetime.strptime(fecha_corte, '%Y-%m-%d').year if fecha_corte else datetime.now().year
                ventas = self._sumar_cuentas_periodo('41', empresa_siglas, anio, 'acreedor')
                costo_ventas = self._sumar_cuentas_periodo('51', empresa_siglas, anio, 'deudor')
                gastos_operacionales = self._sumar_cuentas_periodo('52', empresa_siglas, anio, 'deudor') + \
                                       self._sumar_cuentas_periodo('53', empresa_siglas, anio, 'deudor')
                utilidad_neta = ventas - costo_ventas - gastos_operacionales

                # ========== CALCULAR RATIOS ==========
                ratios = {
                    # LIQUIDEZ
                    'liquidez': {
                        'razon_corriente': self._safe_divide(activo_corriente, pasivo_corriente),
                        'prueba_acida': self._safe_divide(activo_corriente - inventarios, pasivo_corriente),
                        'capital_trabajo': float(activo_corriente - pasivo_corriente),
                        'liquidez_inmediata': self._safe_divide(caja_bancos, pasivo_corriente),
                    },
                    # ENDEUDAMIENTO
                    'endeudamiento': {
                        'endeudamiento_total': self._safe_divide(total_pasivo, total_activo) * 100,
                        'endeudamiento_patrimonio': self._safe_divide(total_pasivo, total_patrimonio) * 100,
                        'apalancamiento': self._safe_divide(total_activo, total_patrimonio),
                        'concentracion_corto_plazo': self._safe_divide(pasivo_corriente, total_pasivo) * 100,
                    },
                    # RENTABILIDAD
                    'rentabilidad': {
                        'margen_bruto': self._safe_divide(ventas - costo_ventas, ventas) * 100,
                        'margen_operacional': self._safe_divide(ventas - costo_ventas - gastos_operacionales,
                                                                ventas) * 100,
                        'margen_neto': self._safe_divide(utilidad_neta, ventas) * 100,
                        'roa': self._safe_divide(utilidad_neta, total_activo) * 100,
                        'roe': self._safe_divide(utilidad_neta, total_patrimonio) * 100,
                    },
                    # ACTIVIDAD
                    'actividad': {
                        'rotacion_activos': self._safe_divide(ventas, total_activo),
                        'rotacion_inventarios': self._safe_divide(costo_ventas, inventarios),
                        'dias_inventario': self._safe_divide(365, self._safe_divide(costo_ventas,
                                                                                    inventarios)) if inventarios > 0 else 0,
                        'rotacion_cuentas_cobrar': self._safe_divide(ventas, cuentas_cobrar),
                        'periodo_cobro': self._safe_divide(365, self._safe_divide(ventas,
                                                                                  cuentas_cobrar)) if cuentas_cobrar > 0 else 0,
                    },
                    # VALORES BASE
                    'valores': {
                        'activo_corriente': format(activo_corriente, '.2f'),
                        'activo_no_corriente': format(activo_no_corriente, '.2f'),
                        'total_activo': format(total_activo, '.2f'),
                        'pasivo_corriente': format(pasivo_corriente, '.2f'),
                        'pasivo_no_corriente': format(pasivo_no_corriente, '.2f'),
                        'total_pasivo': format(total_pasivo, '.2f'),
                        'total_patrimonio': format(total_patrimonio, '.2f'),
                        'ventas': format(ventas, '.2f'),
                        'utilidad_neta': format(utilidad_neta, '.2f'),
                    }
                }

                # Formatear ratios a 2 decimales
                for categoria, valores in ratios.items():
                    if categoria != 'valores':
                        for key, value in valores.items():
                            ratios[categoria][key] = round(value, 2)

                data = ratios

            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def _safe_divide(self, numerador, denominador):
        if denominador == 0:
            return 0
        return float(numerador) / float(denominador)

    def _sumar_cuentas(self, prefijo, empresa_siglas, fecha_corte, naturaleza):
        filtros = {
            'cuenta__empresa__siglas': empresa_siglas,
            'cuenta__codigo__startswith': prefijo,
            'cuenta__tipo_cuenta': 'DETALLE',
            'encabezadocuentaplan__reg_control': 'RT'
        }
        if fecha_corte:
            filtros['encabezadocuentaplan__fecha__lte'] = fecha_corte

        totales = DetalleCuentasPlanCuenta.objects.filter(**filtros).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        if naturaleza == 'deudor':
            return totales['debe'] - totales['haber']
        else:
            return totales['haber'] - totales['debe']

    def _sumar_cuentas_periodo(self, prefijo, empresa_siglas, anio, naturaleza):
        filtros = {
            'cuenta__empresa__siglas': empresa_siglas,
            'cuenta__codigo__startswith': prefijo,
            'cuenta__tipo_cuenta': 'DETALLE',
            'encabezadocuentaplan__reg_control': 'RT',
            'encabezadocuentaplan__fecha__year': anio
        }

        totales = DetalleCuentasPlanCuenta.objects.filter(**filtros).aggregate(
            debe=Coalesce(Sum('debe'), Decimal('0')),
            haber=Coalesce(Sum('haber'), Decimal('0'))
        )

        if naturaleza == 'deudor':
            return abs(totales['debe'] - totales['haber'])
        else:
            return abs(totales['haber'] - totales['debe'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Ratios Financieros'
        context['title'] = 'Indicadores Financieros'
        return context