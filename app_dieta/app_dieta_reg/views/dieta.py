# Create your views here.
import json
import datetime
import decimal
from decimal import Decimal, ROUND_HALF_UP
from django.db import transaction
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, View, UpdateView, DeleteView, TemplateView
from weasyprint import HTML
from app_dieta.app_dieta_reg.forms import AnioDietaForm, RegistroDiaDietaForm, DiaDietaForm, DescripcionDietaForm
from app_dieta.app_dieta_reg.models import MesDieta, AnioDieta, DiaDietaRegistro, DetalleDiaDieta, DescripcionDieta
from django.http import HttpResponse, Http404, JsonResponse, HttpResponseRedirect
# Para crear las Dietas
from app_empresa.app_reg_empresa.models import Empresa, Piscinas
from app_inventario.app_categoria.models import Producto
from app_reportes.utils import render_to_pdf
from app_stock.app_detalle_stock.models import Producto_Stock, Total_Stock
from crum import get_current_user
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app_stock.app_detalle_stock.services import eliminar_asientos_por_detalle
from app_stock.app_detalle_stock.services import revertir_stock_por_detalle


class crearAnioDietaView(CreateView):
    model = AnioDieta
    form_class = AnioDietaForm
    template_name = 'app_dieta/dieta_principal_anio_crear.html'
    success_url = reverse_lazy('app_dieta:principal_anio')

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Dieta'
        return context


class crearMesDietaView(CreateView):
    model = MesDieta
    template_name = 'app_dieta/dieta_principal_mes_crear.html'
    success_url = reverse_lazy('app_dieta:principal_mes')

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Dieta'
        return context


class crearDiaDietaView(CreateView):
    model = DetalleDiaDieta
    form_class = DiaDietaForm
    template_name = 'app_dieta/app_dias_dietas/frm_dieta_dia_cuerpo.html'
    success_url = reverse_lazy('app_dieta:principal_dia')
    url_redirect = success_url

    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_piscinas':
                data = []
                empresa = request.POST['empresa']
                queryset = Piscinas.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                queryset = queryset.filter(empresa__siglas=empresa, prec__exact=False).exclude(id__in=ids_exclude)
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'search_balanceado':
                print('llego aqui a buscar balanceado')
                data = []
                queryset = Producto.objects.all()
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'create':

                with transaction.atomic():

                    factura = DiaDietaRegistro.objects.get(id=self.kwargs['pk'])
                    items = json.loads(request.POST.get('items', '[]'))
                    factura.mes_dieta_id = factura.mes_dieta.pk
                    factura.fecha = request.POST.get('fecha')
                    factura.tip_dieta = True
                    factura.save()

                    for i in items:
                        inv = DetalleDiaDieta()
                        inv.dieta_id = factura.id
                        inv.piscinas_id = int(i['id']) if i.get('id') else None
                        # Balanceado
                        inv.balanceado_id = i.get('balanceado') or None
                        inv.cantidad = Decimal(i['cantidad']) if i.get('cantidad') else Decimal('0')
                        # Insumo 1
                        inv.insumo1 = int(i['insumo1']) if i.get('insumo1') else 0
                        inv.gramaje1 = Decimal(i['gramaje1']) if i.get('gramaje1') else Decimal('0')
                        # Insumo 2
                        inv.insumo2 = int(i['insumo2']) if i.get('insumo2') else 0
                        inv.gramaje2 = Decimal(i['gramaje2']) if i.get('gramaje2') else Decimal('0')
                        # Insumo 3
                        inv.insumo3 = int(i['insumo3']) if i.get('insumo3') else 0
                        inv.gramaje3 = Decimal(i['gramaje3']) if i.get('gramaje3') else Decimal('0')
                        # Insumo 4
                        inv.insumo4 = int(i['insumo4']) if i.get('insumo4') else 0
                        inv.gramaje4 = Decimal(i['gramaje4']) if i.get('gramaje4') else Decimal('0')
                        inv.save()
                    data['success'] = True

            elif action == 'upload_excel':
                print('LLEGO A UPLOAD EXCELL Y EMPEZO A RECORRER EL PYTHON DESDE AJAX')
                try:
                    def safe_float(value):
                        try:
                            if value in [None, '-', '']:
                                return 0.0
                            return float(str(value).replace(',', '.'))
                        except:
                            return 0.0

                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]

                    with transaction.atomic():
                        factura = DiaDietaRegistro.objects.get(id=self.kwargs['pk'])
                        factura.mes_dieta_id = factura.mes_dieta.pk
                        factura.fecha = request.POST.get('fecha')
                        factura.tip_dieta = True
                        factura.save()

                        for row in range(3, excel.max_row + 1):
                            orden = excel.cell(row=row, column=1).value
                            if not orden:
                                continue

                            print(f"Procesando fila del Excell: {row} - Orden de Piscina: {orden}")
                            inv = DetalleDiaDieta(dieta_id=factura.pk)

                            # Buscar piscina
                            piscina = Piscinas.objects.filter(orden=orden).first()
                            if piscina:
                                inv.piscinas_id = piscina.id
                                print(f"Piscina encontrada: {piscina}")
                            else:
                                print(f"Piscina con orden {orden} no encontrada")

                            # Balanceado
                            name_balanceado = excel.cell(row=row, column=3).value
                            if name_balanceado and Producto.objects.filter(nombre__exact=name_balanceado).exists():
                                balanceado = Producto.objects.get(nombre__exact=name_balanceado)
                                inv.balanceado_id = balanceado.id
                                inv.cantidad = safe_float(excel.cell(row=row, column=4).value)
                                print(f"Balanceado: {balanceado.nombre} ({inv.cantidad} lb)")
                            else:
                                print(f"Balanceado no encontrado: {name_balanceado}")
                                inv.cantidad = 0

                            # INSUMOS (Desde el 1 hasta el 4to insumo)
                            insumo_cols = [(6, 7), (9, 10), (12, 13), (15, 16)]
                            for idx, (col_name, col_cant) in enumerate(insumo_cols, start=1):
                                name_insumo = excel.cell(row=row, column=col_name).value
                                cant_insumo = safe_float(excel.cell(row=row, column=col_cant).value)
                                if name_insumo and name_insumo not in ['-', '', None]:
                                    insumo = Producto.objects.filter(nombre__exact=name_insumo).first()
                                    if insumo:
                                        setattr(inv, f"insumo{idx}", insumo.id)
                                        setattr(inv, f"gramaje{idx}", cant_insumo)
                                        print(f" Insumo {idx}: {name_insumo} ({cant_insumo} g)")
                                    else:
                                        print(f" Insumo {idx} no encontrado: {name_insumo}")
                                        setattr(inv, f"insumo{idx}", 0)
                                        setattr(inv, f"gramaje{idx}", 0)
                                else:
                                    setattr(inv, f"insumo{idx}", 0)
                                    setattr(inv, f"gramaje{idx}", 0)
                            inv.save()

                    print("Proceso de carga completada sin errores.")
                    data['success'] = True
                except Exception as e:
                    import traceback
                    print("ERROR GENERAL EN UPLOAD EXCEL:", e)
                    traceback.print_exc()
                    data['error'] = str(e)
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            # data['error'] = 'El error es : ' + str(e)
            pass
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).mes_dieta
        context['nombre'] = 'Dia de Dieta - %s %s' % (dieta.mes_dieta, dieta.anio.anio_dieta)
        context['entity'] = 'Registro de Dieta'
        context['list_url'] = self.success_url
        context['action'] = 'create'
        context['piscinas'] = Piscinas.objects.all()
        context['balanceados'] = Producto.objects.filter(categoria__nombre__icontains='BALANCEADOS')
        context['insumos'] = Producto.objects.filter(categoria__nombre__icontains='INSUMOS')
        context['dieta2'] = DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk'])
        context['dieta_registros'] = DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk'])
        context['mes'] = dieta.mes_dieta
        context['prin_dia'] = dieta.id
        context['pk'] = self.kwargs['pk']
        context['det'] = []
        return context


class editarDiaDietaView(UpdateView):
    model = DiaDietaRegistro
    form_class = DiaDietaForm
    template_name = 'app_dieta/app_dias_dietas/frm_dieta_dia_cuerpo.html'
    success_url = reverse_lazy('app_dieta:principal_dia')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_piscinas':
                data = []
                empresa = request.POST['empresa']
                queryset = Piscinas.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                queryset = queryset.filter(empresa__siglas=empresa, prec__exact=False).exclude(id__in=ids_exclude)
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'search_balanceado':
                print('llego aqui a buscar balanceado')
                data = []
                queryset = Producto.objects.all()
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'edit':
                with transaction.atomic():

                    items = json.loads(request.POST['items'])
                    factura = self.get_object()
                    factura.mes_dieta_id = factura.mes_dieta.pk
                    factura.fecha = request.POST['fecha']
                    factura.tip_dieta = True
                    factura.save()

                    # 🔁 REVERTIR STOCK + ELIMINAR DETALLES
                    for s in factura.detallediadieta_set.all():
                        eliminar_asientos_por_detalle(s.pk)

                        revertir_stock_por_detalle(
                            detalle=s,
                            texto_guia='EDICION DE DIETA Y REAJUSTE DE STOCK'
                        )

                        s.delete()

                    # ➕ CREAR NUEVOS DETALLES
                    for i in items:
                        DetalleDiaDieta.objects.create(
                            dieta_id=factura.pk,
                            piscinas_id=i.get('id'),
                            balanceado_id=i.get('balanceado'),
                            cantidad=decimal.Decimal(i['cantidad']) if i.get('balanceado') else 0,
                            insumo1=int(i['insumo1']) if i.get('insumo1') else 0,
                            gramaje1=decimal.Decimal(i['gramaje1']) if i.get('insumo1') else 0,
                            insumo2=int(i['insumo2']) if i.get('insumo2') else 0,
                            gramaje2=decimal.Decimal(i['gramaje2']) if i.get('insumo2') else 0,
                            insumo3=int(i['insumo3']) if i.get('insumo3') else 0,
                            gramaje3=decimal.Decimal(i['gramaje3']) if i.get('insumo3') else 0,
                            insumo4=int(i['insumo4']) if i.get('insumo4') else 0,
                            gramaje4=decimal.Decimal(i['gramaje4']) if i.get('insumo4') else 0,
                        )

                    data['success'] = True

            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_detalle(self):
        data = []
        for i in DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk']):
            item = i.piscinas.toJSON()
            item['balanceado'] = i.balanceado.id if i.balanceado else None
            item['cantidad'] = format(i.cantidad, '.0f')
            item['insumo1'] = format(i.insumo1, '.0f')
            item['gramaje1'] = format(i.gramaje1, '.0f')
            item['insumo2'] = format(i.insumo2, '.0f')
            item['gramaje2'] = format(i.gramaje2, '.0f')
            item['insumo3'] = format(i.insumo3, '.0f')
            item['gramaje3'] = format(i.gramaje3, '.0f')
            item['insumo4'] = format(i.insumo4, '.0f')
            item['gramaje4'] = format(i.gramaje4, '.0f')
            data.append(item)
        return json.dumps(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).mes_dieta
        context['nombre'] = 'Dia de Dieta - %s %s' % (dieta.mes_dieta, dieta.anio.anio_dieta)
        context['entity'] = 'Registro de Dieta'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['mes'] = dieta.mes_dieta
        context['prin_dia'] = dieta.id
        context['pk'] = self.kwargs['pk']
        context['fecha'] = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).fecha
        context['tip_dieta'] = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).tip_dieta
        context['det'] = self.get_detalle
        return context


class eliminarDiaDietaView(DeleteView):
    model = DiaDietaRegistro
    template_name = 'app_dieta/app_dias_dietas/eliminar_dieta_dia.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        # Obtener el pk del mes_dieta para redirigir correctamente
        mes_dieta_pk = self.object.mes_dieta.pk
        return reverse('app_dieta:principal_dia', kwargs={'pk': mes_dieta_pk})

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            with transaction.atomic():
                factura = self.get_object()

                # Guardar el pk del mes antes de eliminar
                mes_dieta_pk = factura.mes_dieta.pk

                # Revertir stock + eliminar asientos + eliminar detalles
                for detalle in factura.detallediadieta_set.all():
                    # Eliminar asientos contables relacionados
                    eliminar_asientos_por_detalle(detalle.pk)

                    # Revertir stock
                    revertir_stock_por_detalle(
                        detalle=detalle,
                        texto_guia='ELIMINACION DE DIETA Y REAJUSTE DE STOCK'
                    )

                    # Eliminar el detalle
                    detalle.delete()

                # Eliminar el registro principal
                factura.delete()

                data['success'] = True
                data['redirect_url'] = reverse('app_dieta:principal_dia', kwargs={'pk': mes_dieta_pk})

        except Exception as e:
            data['error'] = 'Error al eliminar: ' + str(e)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = self.object.mes_dieta
        context['nombre'] = 'Dia de Dieta'
        context['entity'] = 'Eliminar Registro de Dieta'
        context['list_url'] = reverse('app_dieta:principal_dia', kwargs={'pk': dieta.pk})
        context['mes'] = dieta.mes_dieta
        context['fecha'] = self.object.fecha
        return context


class crearDiaDietaPrecriaView(CreateView):
    model = DetalleDiaDieta
    form_class = DiaDietaForm
    template_name = 'app_dieta/app_dias_dietas_prec/frm_dieta_prec_dia_cuerpo.html'
    success_url = reverse_lazy('app_dieta:principal_dia_prec')
    url_redirect = success_url

    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_piscinas':
                data = []
                empresa = request.POST['empresa']
                queryset = Piscinas.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                queryset = queryset.filter(empresa__siglas=empresa, prec__exact=True).exclude(id__in=ids_exclude)
                print('queryset')
                print(queryset)
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'search_balanceado':
                print('llego aqui a buscar balanceado')
                data = []
                queryset = Producto.objects.all()
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    items = json.loads(request.POST['items'])
                    factura = DiaDietaRegistro.objects.get(id=self.kwargs['pk'])
                    factura.mes_dieta_id = factura.mes_dieta.pk
                    factura.fecha = request.POST['fecha']
                    factura.tip_dieta = False
                    factura.save()
                    for i in items:
                        inv = DetalleDiaDieta()
                        inv.dieta_id = factura.pk
                        inv.piscinas_id = int(i['id']) if i.get('id') else None
                        balanceado_id = (i['balanceado']) if i.get('balanceado') else None
                        inv.balanceado_id = balanceado_id
                        inv.cantidad = decimal.Decimal(i['cantidad']) if i.get('cantidad') and balanceado_id is not None else 0
                        inv.insumo1 = int(i['insumo1']) if i.get('insumo1') else 0
                        inv.gramaje1 = decimal.Decimal(i['gramaje1']) if i.get('gramaje1') else decimal.Decimal('0.00')
                        inv.insumo2 = int(i['insumo2']) if i.get('insumo2') else 0
                        inv.gramaje2 = decimal.Decimal(i['gramaje2']) if i.get('gramaje2') else decimal.Decimal('0.00')
                        inv.insumo3 = int(i['insumo3']) if i.get('insumo3') else 0
                        inv.gramaje3 = decimal.Decimal(i['gramaje3']) if i.get('gramaje3') else decimal.Decimal('0.00')
                        inv.insumo4 = int(i['insumo4']) if i.get('insumo4') else 0
                        inv.gramaje4 = decimal.Decimal(i['gramaje4']) if i.get('gramaje4') else decimal.Decimal('0.00')
                        inv.save()
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).mes_dieta
        context['nombre'] = 'Dia de Dieta - %s %s' % (dieta.mes_dieta, dieta.anio.anio_dieta)
        context['entity'] = 'Registro de Dieta'
        context['list_url'] = self.success_url
        context['action'] = 'create'
        context['piscinas'] = Piscinas.objects.all()
        context['balanceados'] = Producto.objects.filter(categoria__nombre__icontains='BALANCEADOS')
        context['insumos'] = Producto.objects.filter(categoria__nombre__icontains='INSUMOS')
        context['dieta2'] = DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk'])
        context['dieta_registros'] = DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk'])
        context['mes'] = dieta.mes_dieta
        context['prin_dia'] = dieta.id
        context['pk'] = self.kwargs['pk']
        context['det'] = []
        return context


class editarDiaDietaPrecriaView(UpdateView):
    model = DiaDietaRegistro
    form_class = DiaDietaForm
    template_name = 'app_dieta/app_dias_dietas_prec/frm_dieta_prec_dia_cuerpo.html'
    success_url = reverse_lazy('app_dieta:principal_dia_prec')
    url_redirect = success_url

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST['action']
            if action == 'search_piscinas':
                data = []
                empresa = request.POST['empresa']
                queryset = Piscinas.objects.all()
                ids_exclude = json.loads(request.POST['ids'])
                queryset = queryset.filter(empresa__siglas=empresa, prec__exact=True).exclude(id__in=ids_exclude)
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            elif action == 'search_balanceado':
                print('llego aqui a buscar balanceado')
                data = []
                queryset = Producto.objects.all()
                for i in queryset:
                    item = i.toJSON()
                    data.append(item)
            # elif action == 'edit':
            #     with transaction.atomic():
            #         items = json.loads(request.POST['items'])
            #         factura = self.get_object()
            #         factura.mes_dieta_id = factura.mes_dieta.pk
            #         factura.fecha = request.POST['fecha']
            #         factura.tip_dieta = False
            #         factura.save()
            #         for s in factura.detallediadieta_set.all():
            #             datos = [(s.balanceado.pk, int(s.cantidad)), (s.insumo1, int(s.gramaje1)), (s.insumo2, int(s.gramaje2)), (s.insumo3, int(s.gramaje3)), (s.insumo4, int(s.gramaje4))]
            #             print(datos)
            #             for d in datos:
            #                 if d[0]:
            #                     ps = Total_Stock.objects.get(nombre_empresa_id=s.piscinas.empresa.pk, nombre_prod_id=int(d[0]))
            #                     if ps:
            #                         producto = Producto_Stock()
            #                         producto.producto_empresa_id = ps.pk
            #                         producto.tipo = 'INGRESO'
            #                         producto.piscinas = s.piscinas.numero
            #                         producto.cantidad_ingreso = float(d[1])
            #                         producto.fecha_ingreso = s.dieta.fecha
            #                         producto.numero_guia = 'EDICION DE DIETA Y REAJUSTE DE STOCK'
            #                         producto.responsable_ingreso = get_current_user()
            #                         producto.activo = False
            #                         producto.save()
            #             s.delete()
            #         for i in items:
            #             inv = DetalleDiaDieta()
            #             inv.dieta_id = factura.pk
            #             inv.piscinas_id = (i['id']) if i.get('id') else None
            #             balanceado_id = (i['balanceado']) if i.get('balanceado') else None
            #             inv.balanceado_id = balanceado_id
            #             inv.cantidad = decimal.Decimal(i['cantidad']) if i.get('cantidad') and balanceado_id is not None else 0
            #             insumo1 = int(i['insumo1']) if i.get('insumo1') else 0
            #             inv.insumo1 = insumo1
            #             inv.gramaje1 = decimal.Decimal(i['gramaje1']) if i.get('gramaje1') and insumo1 > 0 else 0
            #             insumo2 = int(i['insumo2']) if i.get('insumo2') else 0
            #             inv.insumo2 = insumo2
            #             inv.gramaje2 = decimal.Decimal(i['gramaje2']) if i.get('gramaje2') and insumo2 > 0 else 0
            #             insumo3 = int(i['insumo3']) if i.get('insumo3') else 0
            #             inv.insumo3 = insumo3
            #             inv.gramaje3 = decimal.Decimal(i['gramaje3']) if i.get('gramaje3') and insumo3 > 0 else 0
            #             insumo4 = int(i['insumo4']) if i.get('insumo4') else 0
            #             inv.insumo4 = insumo4
            #             inv.gramaje4 = decimal.Decimal(i['gramaje4']) if i.get('gramaje4') and insumo4 > 0 else 0
            #             inv.save()
            elif action == 'edit':
                with transaction.atomic():

                    items = json.loads(request.POST['items'])
                    factura = self.get_object()

                    factura.mes_dieta_id = factura.mes_dieta.pk
                    factura.fecha = request.POST['fecha']
                    factura.tip_dieta = False
                    factura.save()

                    # 🔁 REVERTIR STOCK PRECRÍA
                    for s in factura.detallediadieta_set.all():
                        eliminar_asientos_por_detalle(s.pk)
                        revertir_stock_por_detalle(detalle=s, texto_guia='EDICION DE PRECRIA Y REAJUSTE DE STOCK')
                        s.delete()

                    # ➕ CREAR NUEVOS DETALLES
                    for i in items:
                        DetalleDiaDieta.objects.create(
                            dieta_id=factura.pk,
                            piscinas_id=i.get('id'),
                            balanceado_id=i.get('balanceado'),
                            cantidad=decimal.Decimal(i['cantidad']) if i.get('balanceado') else 0,
                            insumo1=int(i['insumo1']) if i.get('insumo1') else 0,
                            gramaje1=decimal.Decimal(i['gramaje1']) if i.get('insumo1') else 0,
                            insumo2=int(i['insumo2']) if i.get('insumo2') else 0,
                            gramaje2=decimal.Decimal(i['gramaje2']) if i.get('insumo2') else 0,
                            insumo3=int(i['insumo3']) if i.get('insumo3') else 0,
                            gramaje3=decimal.Decimal(i['gramaje3']) if i.get('insumo3') else 0,
                            insumo4=int(i['insumo4']) if i.get('insumo4') else 0,
                            gramaje4=decimal.Decimal(i['gramaje4']) if i.get('insumo4') else 0,
                        )

                    data['success'] = True

            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = 'el error es : ' + str(e)
        return JsonResponse(data, safe=False)

    def get_detalle(self):
        data = []
        for i in DetalleDiaDieta.objects.filter(dieta_id=self.kwargs['pk']):
            item = i.piscinas.toJSON()
            item['balanceado'] = i.balanceado.id if i.balanceado else None
            item['cantidad'] = format(i.cantidad, '.0f')
            item['insumo1'] = format(i.insumo1, '.0f')
            item['gramaje1'] = format(i.gramaje1, '.2f')
            item['insumo2'] = format(i.insumo2, '.0f')
            item['gramaje2'] = format(i.gramaje2, '.2f')
            item['insumo3'] = format(i.insumo3, '.0f')
            item['gramaje3'] = format(i.gramaje3, '.2f')
            item['insumo4'] = format(i.insumo4, '.0f')
            item['gramaje4'] = format(i.gramaje4, '.2f')
            data.append(item)
        return json.dumps(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).mes_dieta
        context['nombre'] = 'Dia de Dieta - %s %s' % (dieta.mes_dieta, dieta.anio.anio_dieta)
        context['entity'] = 'Registro de Dieta'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['mes'] = dieta.mes_dieta
        context['prin_dia'] = dieta.id
        context['pk'] = self.kwargs['pk']
        context['fecha'] = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).fecha
        context['tip_dieta'] = DiaDietaRegistro.objects.get(id=self.kwargs['pk']).tip_dieta
        context['det'] = self.get_detalle
        return context


class eliminarDiaDietaPrecriaView(DeleteView):
    model = DiaDietaRegistro
    template_name = 'app_dieta/app_dias_dietas/eliminar_dieta_dia_prec.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        # Obtener el pk del mes_dieta para redirigir correctamente
        mes_dieta_pk = self.object.mes_dieta.pk
        return reverse('app_dieta:principal_dia_prec', kwargs={'pk': mes_dieta_pk})

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            with transaction.atomic():
                factura = self.get_object()

                # Guardar el pk del mes antes de eliminar
                mes_dieta_pk = factura.mes_dieta.pk

                # Revertir stock + eliminar asientos + eliminar detalles
                for detalle in factura.detallediadieta_set.all():
                    # Eliminar asientos contables relacionados
                    eliminar_asientos_por_detalle(detalle.pk)

                    # Revertir stock
                    revertir_stock_por_detalle(
                        detalle=detalle,
                        texto_guia='ELIMINACION DE DIETA PRECRIA Y REAJUSTE DE STOCK'
                    )

                    # Eliminar el detalle
                    detalle.delete()

                # Eliminar el registro principal
                factura.delete()

                data['success'] = True
                data['redirect_url'] = reverse('app_dieta:principal_dia_prec', kwargs={'pk': mes_dieta_pk})

        except Exception as e:
            data['error'] = 'Error al eliminar: ' + str(e)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dieta = self.object.mes_dieta
        context['nombre'] = 'Dia de Dieta'
        context['entity'] = 'Eliminar Registro de Dieta Precria'
        context['list_url'] = reverse('app_dieta:principal_dia_prec', kwargs={'pk': dieta.pk})
        context['mes'] = dieta.mes_dieta
        context['fecha'] = self.object.fecha
        return context


# Para listar las Dietas Año
class listarDietaAnioPrincipalView(ListView):
    model = AnioDieta
    template_name = 'app_dieta/dieta_principal_anio.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            data = AnioDieta.objects.get(pk=request.POST['id']).toJSON()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    # defino el dicionario para enviar variables a mi plantilla
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Ventana Principal Dieta Año'
        context['dieta'] = AnioDieta.objects.all()
        return context


# Para listar las Dietas Precria Año
class listarDietaAnioPrecriaView(ListView):
    model = AnioDieta
    template_name = 'app_dieta/app_dias_dietas_prec/dieta_principal_anio_prec.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            data = AnioDieta.objects.get(pk=request.POST['id']).toJSON()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    # defino el dicionario para enviar variables a mi plantilla
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Ventana Principal Año Dieta Precria'
        context['dieta'] = AnioDieta.objects.all()
        return context


# dietas del mes, crea, lista y modifica
@login_required(login_url="/")
def listarMesDietas(request, anio):
    contexto = {
        'nombre': 'Ventana Principal Dieta Mes',
        'meses': MesDieta.objects.filter(anio_id=anio),
        'anio': AnioDieta.objects.get(id=anio)
    }
    if request.POST:
        if request.GET.get('nuevo'):
            mes = MesDieta(anio_id=anio, mes_dieta=request.POST.get('mes_dieta'),
                           descripcion=request.POST.get('descripcion'))
        else:
            mes = MesDieta.objects.get(id=request.GET.get('mes'))
            mes.descripcion = request.POST.get('descripcion')
        mes.save()
    return render(request, 'app_dieta/dieta_principal_mes.html', contexto)


# dietas del mes, crea, lista y modifica
@login_required(login_url="/")
def listarMesDietasPrecrias(request, anio):
    contexto = {
        'nombre': 'Ventana Principal Dieta Precria Mes',
        'meses': MesDieta.objects.filter(anio_id=anio),
        'anio': AnioDieta.objects.get(id=anio)
    }
    if request.POST:
        if request.GET.get('nuevo'):
            mes = MesDieta(anio_id=anio, mes_dieta=request.POST.get('mes_dieta'),
                           descripcion=request.POST.get('descripcion'))
        else:
            mes = MesDieta.objects.get(id=request.GET.get('mes'))
            mes.descripcion = request.POST.get('descripcion')
        mes.save()
    return render(request, 'app_dieta/app_dias_dietas_prec/dieta_principal_mes_prec.html', contexto)


# Para listar las Dietas del Dia Piscinas
# @login_required(login_url="/")
# def listarDiasDietas(request, pk):
#     dietas = DiaDietaRegistro.objects.filter(mes_dieta_id=pk)
#     mes = MesDieta.objects.get(id=pk)
#     if request.POST:
#         dietasR = DiaDietaRegistro(mes_dieta_id=pk)
#         dietasR.save()
#         return redirect(reverse('app_dieta:crear_dia_dieta', kwargs={'pk': dietasR.pk}))
#     contexto = {
#         'anio_id': mes.anio.id,
#         'mes': mes,
#         'fecha': datetime.datetime.now(),
#         'dietas': dietas,
#         'nombre': 'Ventana Principal Dieta Dia',
#         'detail': 'Dieta',
#         'extension': '.xlsx',
#     }
#     return render(request, 'app_dieta/app_dias_dietas/frm_dieta_dia_encabezado.html', contexto)

# Para listar las Dietas del Dia Piscinas
# Para listar las Dietas del Dia Piscinas

@login_required(login_url="/")
def listarDiasDietas(request, pk):
    dietas = DiaDietaRegistro.objects.filter(mes_dieta_id=pk)
    mes = MesDieta.objects.get(id=pk)

    if request.POST:
        dietasR = DiaDietaRegistro(mes_dieta_id=pk)
        dietasR.save()
        return redirect(reverse('app_dieta:crear_dia_dieta', kwargs={'pk': dietasR.pk}))

    dietas_por_empresa = {}
    empresas = Empresa.objects.all()

    for empresa in empresas:
        dietas_empresa = DiaDietaRegistro.objects.filter(
            mes_dieta_id=pk,
            tip_dieta=True,
            detallediadieta__piscinas__empresa=empresa
        ).distinct().order_by('fecha')

        if dietas_empresa.exists():
            dietas_por_empresa[empresa.nombre] = dietas_empresa

    contexto = {
        'anio_id': mes.anio.id,
        'mes': mes,
        'fecha': datetime.datetime.now(),
        'dietas': dietas,
        'dietas_por_empresa': dietas_por_empresa,
        'nombre': 'Ventana Principal Dieta Dia',
        'detail': 'Dieta',
        'extension': '.xlsx',
    }
    return render(request, 'app_dieta/app_dias_dietas/frm_dieta_dia_encabezado.html', contexto)

# Para listar las Dietas del Dia Precrias
# @login_required(login_url="/")
# def listarDiasDietasPrecrias(request, pk):
#     dietas = DiaDietaRegistro.objects.filter(mes_dieta_id=pk)
#     mes = MesDieta.objects.get(id=pk)
#     if request.POST:
#         dietasR = DiaDietaRegistro(mes_dieta_id=pk)
#         dietasR.save()
#         return redirect(reverse('app_dieta:crear_dia_dieta_prec', kwargs={'pk': dietasR.pk}))
#     contexto = {
#         'anio_id': mes.anio.id,
#         'mes': mes,
#         'fecha': datetime.datetime.now(),
#         'dietas': dietas,
#         'nombre': 'Ventana Principal Dieta Precria Dia',
#         'detail': 'PrecDieta',
#         'extension': '.xlsx',
#     }
#     return render(request, 'app_dieta/app_dias_dietas_prec/frm_dieta_prec_dia_encabezado.html', contexto)


# CON XHTML2PDF


@login_required(login_url="/")
def listarDiasDietasPrecrias(request, pk):
    dietas = DiaDietaRegistro.objects.filter(mes_dieta_id=pk)
    mes = MesDieta.objects.get(id=pk)

    # 🔹 Crear nuevo día de dieta (PRECRÍA)
    if request.POST:
        dietasR = DiaDietaRegistro(mes_dieta_id=pk)
        dietasR.save()
        return redirect(reverse('app_dieta:crear_dia_dieta_prec', kwargs={'pk': dietasR.pk}))

    dietas_por_empresa = {}
    empresas = Empresa.objects.all()

    for empresa in empresas:
        dietas_empresa = DiaDietaRegistro.objects.filter(
            mes_dieta_id=pk,
            tip_dieta=False,  # PRECRÍA
            detallediadieta__piscinas__empresa=empresa,
            detallediadieta__piscinas__prec=True
        ).distinct().order_by('fecha')

        if dietas_empresa.exists():
            dietas_por_empresa[empresa.nombre] = dietas_empresa

    contexto = {
        'anio_id': mes.anio.id,
        'mes': mes,
        'fecha': datetime.datetime.now(),
        'dietas': dietas,
        'dietas_por_empresa': dietas_por_empresa,
        'nombre': 'Ventana Principal Dieta Precría Día',
        'detail': 'PrecDieta',
        'extension': '.xlsx',
    }
    return render(request, 'app_dieta/app_dias_dietas_prec/frm_dieta_prec_dia_encabezado.html', contexto)



# class ListarDietaPDF(View):
#     def get(self, request, *args, **kwargs):
#         if 'pk' in kwargs:
#             dieta = DetalleDiaDieta.objects.filter(dieta_id=kwargs['pk']).order_by('piscinas_id')
#
#             fecha_dieta = ''
#
#             if dieta:
#                 fecha_dieta = dieta[0].dieta.fecha
#
#             # Empresa PSM
#             balanceado = {}
#             insumo = {}
#             acum = {}
#
#             for b in dieta.filter(piscinas__empresa__siglas='PSM'):
#                 if b.balanceado:
#                     nombre_b = b.balanceado.nombre
#                     prod = Producto.objects.get(nombre__icontains=nombre_b).peso_presentacion
#
#                     if nombre_b not in balanceado:
#                         balanceado[nombre_b] = b.cantidad
#                     else:
#                         balanceado[nombre_b] = balanceado[nombre_b] + b.cantidad
#                         acum = format(balanceado[nombre_b] / prod, '.1f')
#
#                 nombre_i = b.insumo1
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     prod = Producto.objects.get(nombre__icontains=nombre_i).peso_presentacion
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje1
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje1
#                         acum = format(insumo[nombre_i] / prod, '.1f')
#
#                 nombre_i = b.insumo2
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     prod = Producto.objects.get(nombre__icontains=nombre_i).peso_presentacion
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje2
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje2
#                         acum = format(insumo[nombre_i] / prod, '.1f')
#
#                 nombre_i = b.insumo3
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     prod = Producto.objects.get(nombre__icontains=nombre_i).peso_presentacion
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje3
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje3
#                         acum = format(insumo[nombre_i] / prod, '.1f')
#
#                 nombre_i = b.insumo4
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     prod = Producto.objects.get(nombre__icontains=nombre_i).peso_presentacion
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje4
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje4
#                         acum = format(insumo[nombre_i] / prod, '.1f')
#
#             resumen_totales = {
#                 'psm': {'balanceado': balanceado, 'insumo': insumo}
#             }
#
#             # Empresa BIO
#             balanceado = {}
#             insumo = {}
#
#             for b in dieta.filter(piscinas__empresa__siglas='BIO'):
#                 if b.balanceado:
#                     nombre_b = b.balanceado.nombre
#
#                     if nombre_b not in balanceado:
#                         balanceado[nombre_b] = b.cantidad
#                     else:
#                         balanceado[nombre_b] = balanceado[nombre_b] + b.cantidad
#
#                 nombre_i = b.insumo1
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje1
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje1
#
#                 nombre_i = b.insumo2
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje2
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje2
#
#                 nombre_i = b.insumo3
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje3
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje3
#
#                 nombre_i = b.insumo4
#                 if nombre_i:
#                     nombre_i = Producto.objects.get(id=nombre_i).nombre
#                     if nombre_i not in insumo:
#                         insumo[nombre_i] = b.gramaje4
#                     else:
#                         insumo[nombre_i] = insumo[nombre_i] + b.gramaje4
#
#             resumen_totales['bio'] = {'balanceado': balanceado, 'insumo': insumo}
#
#             data = {
#                 'insumos': Producto.objects.filter(categoria__nombre__icontains='INSUMOS'),
#                 'dieta_registros': dieta,
#                 'fecha_dieta': fecha_dieta,
#                 'resumen_totales': resumen_totales
#             }
#             pdf = render_to_pdf('app_reportes/printDieta.html', data)
#             return HttpResponse(pdf, content_type='application/pdf')


# CON WEASYPRINT


class ListarDietaPDF(View):

    def get(self, request, *args, **kwargs):

        if 'pk' not in kwargs:
            return HttpResponse(status=404)

        dieta = (
            DetalleDiaDieta.objects
            .filter(dieta_id=kwargs['pk'])
            .select_related(
                'dieta',
                'balanceado',
                'piscinas',
                'piscinas__empresa'
            )
            .order_by('piscinas_id')
        )

        fecha_dieta = dieta.first().dieta.fecha if dieta.exists() else ''
        # =============================================
        # Cargar TODOS los productos una sola vez
        # =============================================
        productos = Producto.objects.in_bulk()
        # =============================================
        # Función auxiliar
        # =============================================
        def agregar_insumo(diccionario, producto_id, gramaje):

            if not producto_id:
                return

            producto = productos.get(int(producto_id))

            if not producto:
                return

            nombre = producto.nombre

            if nombre not in diccionario:
                diccionario[nombre] = gramaje
            else:
                diccionario[nombre] += gramaje

        # =============================================
        # EMPRESA PSM
        # =============================================

        balanceado = {}
        insumo = {}

        for b in dieta.filter(piscinas__empresa__siglas='PSM'):

            # -----------------------------
            # Balanceado
            # -----------------------------

            if b.balanceado:

                nombre = b.balanceado.nombre

                if nombre not in balanceado:
                    balanceado[nombre] = b.cantidad
                else:
                    balanceado[nombre] += b.cantidad

            # -----------------------------
            # Insumos
            # -----------------------------

            agregar_insumo(insumo, b.insumo1, b.gramaje1)
            agregar_insumo(insumo, b.insumo2, b.gramaje2)
            agregar_insumo(insumo, b.insumo3, b.gramaje3)
            agregar_insumo(insumo, b.insumo4, b.gramaje4)

        resumen_totales = {
            'psm': {
                'balanceado': balanceado,
                'insumo': insumo
            }
        }

        # =============================================
        # EMPRESA BIO
        # =============================================

        balanceado = {}
        insumo = {}

        for b in dieta.filter(piscinas__empresa__siglas='BIO'):

            if b.balanceado:

                nombre = b.balanceado.nombre

                if nombre not in balanceado:
                    balanceado[nombre] = b.cantidad
                else:
                    balanceado[nombre] += b.cantidad

            agregar_insumo(insumo, b.insumo1, b.gramaje1)
            agregar_insumo(insumo, b.insumo2, b.gramaje2)
            agregar_insumo(insumo, b.insumo3, b.gramaje3)
            agregar_insumo(insumo, b.insumo4, b.gramaje4)

        resumen_totales['bio'] = {
            'balanceado': balanceado,
            'insumo': insumo
        }

        data = {
            'insumos': Producto.objects.filter(
                categoria__nombre__icontains='INSUMOS'
            ),
            'dieta_registros': dieta,
            'fecha_dieta': fecha_dieta,
            'resumen_totales': resumen_totales
        }

        pdf = render_to_pdf(
            'app_reportes/printDieta.html',
            data
        )

        return HttpResponse(
            pdf,
            content_type='application/pdf'
        )


class printDieta(View):

    def get(self, request, *args, **kwargs):
        if 'pk' in kwargs:
            dieta = DetalleDiaDieta.objects.filter(dieta_id=kwargs['pk']).order_by('piscinas__orden')
            data = {
                'insumos': Producto.objects.filter(categoria__nombre__icontains='INSUMOS'),
                'dieta_registros': dieta,
            }
            template = get_template("app_reportes/printDieta.html")
            html_template = template.render(data)
            HTML(string=html_template).write_pdf(target="dieta.pdf")


class listarDescripcionDietaView(ListView):
    model = DescripcionDieta
    template_name = 'app_dieta/app_descripcion/listar_descripcion.html'

    @method_decorator(csrf_exempt)
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            data = DescripcionDieta.objects.get(pk=request.POST['id']).toJSON()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    # defino el dicionario para enviar variables a mi plantilla
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Descripción de Escaneo de Dietas'
        context['descripcion_dieta'] = DescripcionDieta.objects.all()
        return context


class crearDescripcionDietaView(CreateView):
    model = DescripcionDieta
    form_class = DescripcionDietaForm
    template_name = 'app_dieta/app_descripcion/crear_descripcion.html'
    success_url = reverse_lazy('app_dieta:listar_descripcion_dieta')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Descripción de Dieta'
        return context


class actualizarDescripcionDietaView(UpdateView):
    model = DescripcionDieta
    form_class = DescripcionDietaForm
    template_name = 'app_dieta/app_descripcion/crear_descripcion.html'
    success_url = reverse_lazy('app_dieta:listar_descripcion_dieta')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Descripción de Dieta'
        context['action'] = 'crear'
        return context


class eliminarDescripcionDietaView(DeleteView):
    model = DescripcionDieta
    form_class = DescripcionDietaForm
    template_name = 'app_dieta/app_descripcion/eliminar_descripcion.html'
    success_url = reverse_lazy('app_dieta:listar_descripcion_dieta')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nombre'] = 'Descripción de Dieta'
        context['action'] = 'crear'
        return context


class ReporteDietaDiaView(TemplateView):
    template_name = 'app_dieta/reportes/reporte_dieta_dia.html'

    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['empresas'] = Empresa.objects.all()
        context['titulo'] = 'Reporte de Dietas por Día'
        return context

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action', '')

            if action == 'filtrar':
                fecha_desde = request.POST.get('fecha_desde', '')
                fecha_hasta = request.POST.get('fecha_hasta', '')
                empresa_id = request.POST.get('empresa', '')

                # Filtrar DiaDietaRegistro por rango de fechas
                queryset = DiaDietaRegistro.objects.all()

                if fecha_desde:
                    queryset = queryset.filter(fecha__gte=fecha_desde)
                if fecha_hasta:
                    queryset = queryset.filter(fecha__lte=fecha_hasta)

                # Obtener todos los productos para mapear IDs a nombres
                productos_dict = {p.id: p.nombre for p in Producto.objects.all()}

                # Estructura de datos agrupada por fecha
                dietas_por_dia = []

                for dieta_registro in queryset.order_by('fecha'):
                    # Filtrar detalles por empresa si se especificó
                    detalles = DetalleDiaDieta.objects.filter(dieta=dieta_registro)

                    if empresa_id:
                        detalles = detalles.filter(piscinas__empresa_id=empresa_id)

                    if not detalles.exists():
                        continue

                    filas = []
                    for idx, det in enumerate(detalles.select_related('piscinas', 'balanceado'), 1):
                        fila = {
                            'numero': idx,
                            'piscina': det.piscinas.numero if det.piscinas else '-',
                            'balanceado': det.balanceado.nombre if det.balanceado else '-',
                            'libras': float(det.cantidad) if det.cantidad else 0,
                            'insumo1_nombre': productos_dict.get(det.insumo1, '') if det.insumo1 else '',
                            'insumo1_gramos': float(det.gramaje1) if det.gramaje1 else 0,
                            'insumo2_nombre': productos_dict.get(det.insumo2, '') if det.insumo2 else '',
                            'insumo2_gramos': float(det.gramaje2) if det.gramaje2 else 0,
                            'insumo3_nombre': productos_dict.get(det.insumo3, '') if det.insumo3 else '',
                            'insumo3_gramos': float(det.gramaje3) if det.gramaje3 else 0,
                            'insumo4_nombre': productos_dict.get(det.insumo4, '') if det.insumo4 else '',
                            'insumo4_gramos': float(det.gramaje4) if det.gramaje4 else 0,
                        }
                        filas.append(fila)

                    if filas:
                        dietas_por_dia.append({
                            'fecha': dieta_registro.fecha.strftime('%d-%m-%Y') if dieta_registro.fecha else '-',
                            'dieta_id': dieta_registro.id,
                            'filas': filas
                        })

                data = {
                    'success': True,
                    'dietas': dietas_por_dia,
                    'fecha_desde': fecha_desde or 'Inicio',
                    'fecha_hasta': fecha_hasta or 'Hoy',
                    'empresa_nombre': Empresa.objects.get(id=empresa_id).nombre if empresa_id else 'TODAS LAS EMPRESAS'
                }

        except Exception as e:
            data = {'error': str(e)}

        return JsonResponse(data, safe=False)




# class CopiarGuardarView(TemplateView):
#     template_name = 'app_copiarguardar/copiar_pegar.html'
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#
#         try:
#             action = request.POST.get('action')
#
#             if action == 'edit':
#
#                 with transaction.atomic():
#
#                     items = json.loads(request.POST['items'])
#                     fecha = request.POST.get('fecha')
#
#                     # 🔥 CREAR CABECERA (puedes adaptar a tu modelo real)
#                     dieta = DiaDietaRegistro.objects.create(
#                         fecha=fecha,
#                         tip_dieta=False
#                     )
#
#                     # 🔥 CREAR DETALLES (AQUÍ ESTÁ LO IMPORTANTE)
#                     for i in items:
#
#                         DetalleDiaDieta.objects.create(
#                             dieta_id=dieta.pk,
#                             piscinas_id=i.get('id'),
#
#                             balanceado_id=i.get('balanceado'),
#
#                             cantidad=decimal.Decimal(i.get('cantidad', 0)) if i.get('balanceado') else 0,
#
#                             insumo1=int(i.get('insumo1', 0)),
#                             gramaje1=decimal.Decimal(i.get('gramaje1', 0)) if i.get('insumo1') else 0,
#
#                             insumo2=int(i.get('insumo2', 0)),
#                             gramaje2=decimal.Decimal(i.get('gramaje2', 0)) if i.get('insumo2') else 0,
#
#                             insumo3=int(i.get('insumo3', 0)),
#                             gramaje3=decimal.Decimal(i.get('gramaje3', 0)) if i.get('insumo3') else 0,
#
#                             insumo4=int(i.get('insumo4', 0)),
#                             gramaje4=decimal.Decimal(i.get('gramaje4', 0)) if i.get('insumo4') else 0,
#                         )
#
#                     data['success'] = True
#
#             else:
#                 data['error'] = 'Acción no válida'
#
#         except Exception as e:
#             data['error'] = str(e)
#
#         return JsonResponse(data, safe=False)



class CopiarGuardarView(TemplateView):
    template_name = 'app_copiarguardar/copiar_pegar.html'

    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'edit':
                print('LLEGO A COPIAR/PEGAR Y EMPEZO A RECORRER EL PYTHON DESDE AJAX')

                try:
                    def safe_decimal(value):
                        try:
                            if value in [None, '-', '', 0, '0']:
                                return Decimal('0')
                            return Decimal(str(value).replace(',', '.'))
                        except:
                            return Decimal('0')

                    def buscar_producto(nombre):
                        nombre = (nombre or '').strip()
                        if not nombre or nombre in ['-', '0']:
                            return None
                        prod = Producto.objects.filter(nombre__iexact=nombre).first()
                        if not prod:
                            prod = Producto.objects.filter(nombre__icontains=nombre).first()
                        return prod

                    items = json.loads(request.POST.get('items', '[]'))
                    fecha_str = request.POST.get('fecha')

                    if not items:
                        data['error'] = 'No hay datos para guardar'
                        return JsonResponse(data, safe=False)

                    if not fecha_str:
                        data['error'] = 'Debe seleccionar una fecha'
                        return JsonResponse(data, safe=False)

                    # FIX 1: usar .date() para que SIEMPRE coincida con el DateField guardado
                    fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()

                    # FIX 2: eliminar filas repetidas del Excel (misma piscina) -> se queda la ultima
                    filas_unicas = {}
                    for i in items:
                        pid = i.get('id')
                        if pid is None or str(pid).strip() == '':
                            continue
                        filas_unicas[str(pid)] = i
                    items = list(filas_unicas.values())

                    meses = {
                        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
                        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
                        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
                    }

                    with transaction.atomic():
                        anio_obj, _ = AnioDieta.objects.get_or_create(anio_dieta=fecha_obj.year)
                        mes_obj, _ = MesDieta.objects.get_or_create(anio=anio_obj, mes_dieta=meses[fecha_obj.month])

                        # ---- Agrupar items por empresa ----
                        grupos_por_empresa = {}
                        errores = []

                        for i in items:
                            piscina_orden = i.get('id')
                            if not piscina_orden:
                                continue

                            piscina = Piscinas.objects.select_related('empresa').filter(orden=piscina_orden).first()
                            if not piscina:
                                errores.append(f"Piscina orden {piscina_orden} no encontrada")
                                continue
                            if not piscina.empresa:
                                errores.append(f"Piscina orden {piscina_orden} sin empresa asignada")
                                continue

                            emp = piscina.empresa
                            grupo = grupos_por_empresa.setdefault(emp.id, {'empresa': emp, 'filas': []})
                            grupo['filas'].append((i, piscina))

                        guardados = 0

                        # ---- Procesar cada empresa ----
                        for emp_id, grupo in grupos_por_empresa.items():
                            empresa = grupo['empresa']

                            # FIX 3: traer TODOS los registros existentes (no .first()) y BLOQUEARLOS
                            #        con select_for_update para evitar doble envio simultaneo.
                            # Obtener primero las dietas que pertenecen a la empresa
                            ids_dietas = (
                                DetalleDiaDieta.objects
                                    .filter(piscinas__empresa=empresa)
                                    .values_list('dieta_id', flat=True)
                                    .distinct()
                            )

                            # Bloquear únicamente los registros necesarios
                            print("ANTES DE BUSCAR FACTURAS")
                            facturas_existentes = list(
                                DiaDietaRegistro.objects
                                    .select_for_update()
                                    .filter(
                                    id__in=ids_dietas,
                                    mes_dieta=mes_obj,
                                    fecha=fecha_obj
                                )
                            )
                            print("DESPUES DE BUSCAR FACTURAS")

                            if facturas_existentes:
                                factura = facturas_existentes[0]  # conservamos el primero
                                # Reversar stock de TODOS los detalles y eliminar registros sobrantes
                                for f in facturas_existentes:
                                    for s in f.detallediadieta_set.all():
                                        s.delete()  # dispara reversa de stock/kardex
                                    if f.id != factura.id:
                                        f.delete()  # FIX: elimina duplicados previos ya existentes
                                print(f"Registros anteriores limpiados para {empresa} en {fecha_obj}")
                            else:
                                factura = DiaDietaRegistro.objects.create(
                                    mes_dieta=mes_obj,
                                    fecha=fecha_obj,
                                    tip_dieta=True
                                )
                                print(f"Nuevo registro creado para empresa {empresa} en {fecha_obj}")

                            factura.tip_dieta = True
                            factura.save()

                            # ---- Crear detalles ----
                            for i, piscina in grupo['filas']:
                                inv = DetalleDiaDieta(dieta_id=factura.pk)
                                inv.piscinas_id = piscina.id

                                balanceado = buscar_producto(i.get('balanceado', ''))
                                if balanceado:
                                    inv.balanceado_id = balanceado.id
                                inv.cantidad = safe_decimal(i.get('cantidad', 0))

                                for num in range(1, 5):
                                    insumo = buscar_producto(i.get(f'insumo{num}', ''))
                                    if insumo:
                                        setattr(inv, f"insumo{num}", insumo.id)
                                        setattr(inv, f"gramaje{num}", safe_decimal(i.get(f'gramaje{num}', 0)))
                                    else:
                                        setattr(inv, f"insumo{num}", 0)
                                        setattr(inv, f"gramaje{num}", Decimal('0'))

                                inv.save()  # genera los EGRESOS de stock
                                guardados += 1

                        print(f"Proceso completado. Guardados: {guardados}, Errores: {len(errores)}")

                        data['success'] = True
                        data['guardados'] = guardados
                        data['registros_empresas'] = len(grupos_por_empresa)
                        data['errores'] = errores

                except Exception as e:
                    import traceback
                    print("ERROR GENERAL EN COPIAR/PEGAR:", e)
                    traceback.print_exc()
                    data['error'] = str(e)

            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            import traceback
            print("ERROR:", e)
            traceback.print_exc()
            data['error'] = str(e)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Copiar Dieta desde Excel'
        return context



# class CopiarGuardarPrecriaView(TemplateView):
#     template_name = 'app_copiarguardar/copiar_pegar_prec.html'
#
#     @method_decorator(login_required)
#     @method_decorator(csrf_exempt)
#     def dispatch(self, request, *args, **kwargs):
#         return super().dispatch(request, *args, **kwargs)
#
#     def post(self, request, *args, **kwargs):
#         data = {}
#         try:
#             action = request.POST.get('action')
#
#             if action == 'edit':
#                 print('LLEGO A COPIAR/PEGAR PRECRIA Y EMPEZO A RECORRER EL PYTHON DESDE AJAX')
#
#                 try:
#                     def safe_decimal(value):
#                         try:
#                             if value in [None, '-', '', 0, '0']:
#                                 return Decimal('0')
#                             return Decimal(str(value).replace(',', '.'))
#                         except:
#                             return Decimal('0')
#
#                     def buscar_producto(nombre):
#                         nombre = (nombre or '').strip()
#                         if not nombre or nombre in ['-', '0']:
#                             return None
#                         prod = Producto.objects.filter(nombre__iexact=nombre).first()
#                         if not prod:
#                             prod = Producto.objects.filter(nombre__icontains=nombre).first()
#                         return prod
#
#                     items = json.loads(request.POST.get('items', '[]'))
#                     fecha_str = request.POST.get('fecha')
#
#                     if not items:
#                         data['error'] = 'No hay datos para guardar'
#                         return JsonResponse(data, safe=False)
#
#                     if not fecha_str:
#                         data['error'] = 'Debe seleccionar una fecha'
#                         return JsonResponse(data, safe=False)
#
#                     # FIX 1: usar .date() para que SIEMPRE coincida con el DateField guardado
#                     fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
#
#                     # FIX 2: eliminar filas repetidas del Excel (misma piscina) -> se queda la ultima
#                     filas_unicas = {}
#                     for i in items:
#                         pid = i.get('id')
#                         if pid is None or str(pid).strip() == '':
#                             continue
#                         filas_unicas[str(pid)] = i
#                     items = list(filas_unicas.values())
#
#                     meses = {
#                         1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
#                         5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
#                         9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
#                     }
#
#                     with transaction.atomic():
#                         anio_obj, _ = AnioDieta.objects.get_or_create(anio_dieta=fecha_obj.year)
#                         mes_obj, _ = MesDieta.objects.get_or_create(anio=anio_obj, mes_dieta=meses[fecha_obj.month])
#
#                         # ---- Agrupar items por empresa ----
#                         grupos_por_empresa = {}
#                         errores = []
#
#                         for i in items:
#                             piscina_orden = i.get('id')
#                             if not piscina_orden:
#                                 continue
#
#                             # PRECRIA: solo piscinas de precria (prec=True)
#                             piscina = Piscinas.objects.select_related('empresa').filter(
#                                 orden=piscina_orden, prec__exact=True
#                             ).first()
#                             if not piscina:
#                                 errores.append(f"Piscina orden {piscina_orden} no encontrada (precria)")
#                                 continue
#                             if not piscina.empresa:
#                                 errores.append(f"Piscina orden {piscina_orden} sin empresa asignada")
#                                 continue
#
#                             emp = piscina.empresa
#                             grupo = grupos_por_empresa.setdefault(emp.id, {'empresa': emp, 'filas': []})
#                             grupo['filas'].append((i, piscina))
#
#                         guardados = 0
#
#                         # ---- Procesar cada empresa ----
#                         for emp_id, grupo in grupos_por_empresa.items():
#                             empresa = grupo['empresa']
#
#                             # FIX 3: traer TODOS los registros existentes (no .first()) y BLOQUEARLOS
#                             #        con select_for_update para evitar doble envio simultaneo.
#                             # Obtener primero las precrias que pertenecen a la empresa
#                             ids_dietas = (
#                                 DetalleDiaDieta.objects
#                                     .filter(piscinas__empresa=empresa)
#                                     .values_list('dieta_id', flat=True)
#                                     .distinct()
#                             )
#
#                             # Bloquear unicamente los registros necesarios
#                             # PRECRIA: tip_dieta = False
#                             print("ANTES DE BUSCAR FACTURAS (PRECRIA)")
#                             facturas_existentes = list(
#                                 DiaDietaRegistro.objects
#                                     .select_for_update()
#                                     .filter(
#                                     id__in=ids_dietas,
#                                     mes_dieta=mes_obj,
#                                     fecha=fecha_obj,
#                                     tip_dieta=False
#                                 )
#                             )
#                             print("DESPUES DE BUSCAR FACTURAS (PRECRIA)")
#
#                             if facturas_existentes:
#                                 factura = facturas_existentes[0]  # conservamos el primero
#                                 # Reversar stock de TODOS los detalles y eliminar registros sobrantes
#                                 for f in facturas_existentes:
#                                     for s in f.detallediadieta_set.all():
#                                         s.delete()  # dispara reversa de stock/kardex
#                                     if f.id != factura.id:
#                                         f.delete()  # FIX: elimina duplicados previos ya existentes
#                                 print(f"Registros anteriores de precria limpiados para {empresa} en {fecha_obj}")
#                             else:
#                                 factura = DiaDietaRegistro.objects.create(
#                                     mes_dieta=mes_obj,
#                                     fecha=fecha_obj,
#                                     tip_dieta=False
#                                 )
#                                 print(f"Nuevo registro de precria creado para empresa {empresa} en {fecha_obj}")
#
#                             factura.tip_dieta = False
#                             factura.save()
#
#                             # ---- Crear detalles ----
#                             for i, piscina in grupo['filas']:
#                                 inv = DetalleDiaDieta(dieta_id=factura.pk)
#                                 inv.piscinas_id = piscina.id
#
#                                 balanceado = buscar_producto(i.get('balanceado', ''))
#                                 if balanceado:
#                                     inv.balanceado_id = balanceado.id
#                                 inv.cantidad = safe_decimal(i.get('cantidad', 0))
#
#                                 for num in range(1, 5):
#                                     insumo = buscar_producto(i.get(f'insumo{num}', ''))
#                                     if insumo:
#                                         setattr(inv, f"insumo{num}", insumo.id)
#                                         setattr(inv, f"gramaje{num}", safe_decimal(i.get(f'gramaje{num}', 0)))
#                                     else:
#                                         setattr(inv, f"insumo{num}", 0)
#                                         setattr(inv, f"gramaje{num}", Decimal('0'))
#
#                                 inv.save()  # genera los EGRESOS de stock
#                                 guardados += 1
#
#                         print(f"Proceso de precria completado. Guardados: {guardados}, Errores: {len(errores)}")
#
#                         data['success'] = True
#                         data['guardados'] = guardados
#                         data['registros_empresas'] = len(grupos_por_empresa)
#                         data['errores'] = errores
#
#                 except Exception as e:
#                     import traceback
#                     print("ERROR GENERAL EN COPIAR/PEGAR PRECRIA:", e)
#                     traceback.print_exc()
#                     data['error'] = str(e)
#
#             else:
#                 data['error'] = 'Accion no valida'
#
#         except Exception as e:
#             import traceback
#             print("ERROR:", e)
#             traceback.print_exc()
#             data['error'] = str(e)
#
#         return JsonResponse(data, safe=False)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['title'] = 'Copiar Precria desde Excel'
#         return context



class CopiarGuardarPrecriaView(TemplateView):
    template_name = 'app_copiarguardar/copiar_pegar_prec.html'

    @method_decorator(login_required)
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            action = request.POST.get('action')

            if action == 'edit':
                print('LLEGO A COPIAR/PEGAR PRECRIA Y EMPEZO A RECORRER EL PYTHON DESDE AJAX')

                try:
                    def safe_decimal(value):
                        try:
                            if value in [None, '-', '', 0, '0']:
                                return Decimal('0')
                            return Decimal(str(value).replace(',', '.'))
                        except:
                            return Decimal('0')

                    def buscar_producto(nombre):
                        nombre = (nombre or '').strip()
                        if not nombre or nombre in ['-', '0']:
                            return None
                        prod = Producto.objects.filter(nombre__iexact=nombre).first()
                        if not prod:
                            prod = Producto.objects.filter(nombre__icontains=nombre).first()
                        return prod

                    items = json.loads(request.POST.get('items', '[]'))
                    # PRECRIA: la PISCINA se elige en el formulario (viene su "orden")
                    piscina_orden = request.POST.get('piscina')

                    if not piscina_orden:
                        data['error'] = 'Debe seleccionar una piscina'
                        return JsonResponse(data, safe=False)

                    if not items:
                        data['error'] = 'No hay datos para guardar'
                        return JsonResponse(data, safe=False)

                    # Buscar la piscina seleccionada (solo precria: prec=True)
                    piscina = (
                        Piscinas.objects
                        .select_related('empresa')
                        .filter(orden=piscina_orden, prec__exact=True)
                        .first()
                    )
                    if not piscina:
                        data['error'] = f'Piscina orden {piscina_orden} no encontrada (precria)'
                        return JsonResponse(data, safe=False)
                    if not piscina.empresa:
                        data['error'] = f'Piscina orden {piscina_orden} sin empresa asignada'
                        return JsonResponse(data, safe=False)

                    empresa = piscina.empresa

                    meses = {
                        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
                        5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
                        9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
                    }

                    # FIX: eliminar filas repetidas (misma fecha) -> se queda la ultima
                    filas_unicas = {}
                    for i in items:
                        fkey = (i.get('fecha') or '').strip()
                        if not fkey:
                            continue
                        filas_unicas[fkey] = i
                    items = list(filas_unicas.values())

                    guardados = 0
                    errores = []

                    with transaction.atomic():

                        # PRECRIA: cada FILA es un DIA distinto -> un registro por fecha
                        for i in items:
                            fecha_str = i.get('fecha')
                            if not fecha_str:
                                errores.append('Fila sin fecha, omitida')
                                continue

                            try:
                                # FIX: .date() para que coincida con el DateField guardado
                                fecha_obj = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                            except Exception:
                                errores.append(f'Fecha invalida: {fecha_str}')
                                continue

                            anio_obj, _ = AnioDieta.objects.get_or_create(anio_dieta=fecha_obj.year)
                            mes_obj, _ = MesDieta.objects.get_or_create(
                                anio=anio_obj, mes_dieta=meses[fecha_obj.month]
                            )

                            # Buscar el registro de PRECRIA (tip_dieta=False) de esta empresa
                            # para esta fecha. Se comparte entre las piscinas de precria de la
                            # misma empresa en el mismo dia.
                            ids_dietas = (
                                DetalleDiaDieta.objects
                                .filter(piscinas__empresa=empresa)
                                .values_list('dieta_id', flat=True)
                                .distinct()
                            )

                            # Bloquear con select_for_update para evitar doble envio simultaneo
                            factura = (
                                DiaDietaRegistro.objects
                                .select_for_update()
                                .filter(
                                    id__in=ids_dietas,
                                    mes_dieta=mes_obj,
                                    fecha=fecha_obj,
                                    tip_dieta=False
                                )
                                .first()
                            )

                            if not factura:
                                factura = DiaDietaRegistro.objects.create(
                                    mes_dieta=mes_obj,
                                    fecha=fecha_obj,
                                    tip_dieta=False
                                )
                                print(f"Nuevo registro PRECRIA creado {empresa} {fecha_obj}")

                            factura.tip_dieta = False
                            factura.save()

                            # Reversar/limpiar SOLO el detalle de ESTA piscina en ese dia
                            # (no tocar las demas piscinas del mismo registro)
                            for s in factura.detallediadieta_set.filter(piscinas=piscina):
                                s.delete()  # dispara reversa de stock/kardex

                            # Crear el detalle de la piscina para ese dia
                            inv = DetalleDiaDieta(dieta_id=factura.pk)
                            inv.piscinas_id = piscina.id

                            balanceado = buscar_producto(i.get('balanceado', ''))
                            if balanceado:
                                inv.balanceado_id = balanceado.id
                                inv.cantidad = safe_decimal(i.get('cantidad', 0))
                            else:
                                inv.cantidad = Decimal('0')

                            for num in range(1, 5):
                                insumo = buscar_producto(i.get(f'insumo{num}', ''))
                                if insumo:
                                    setattr(inv, f"insumo{num}", insumo.id)
                                    setattr(inv, f"gramaje{num}", safe_decimal(i.get(f'gramaje{num}', 0)))
                                else:
                                    setattr(inv, f"insumo{num}", 0)
                                    setattr(inv, f"gramaje{num}", Decimal('0'))

                            inv.save()  # genera los EGRESOS de stock
                            guardados += 1

                        print(f"Proceso PRECRIA completado. Guardados: {guardados}, Errores: {len(errores)}")

                        data['success'] = True
                        data['guardados'] = guardados
                        data['errores'] = errores

                except Exception as e:
                    import traceback
                    print("ERROR GENERAL EN COPIAR/PEGAR PRECRIA:", e)
                    traceback.print_exc()
                    data['error'] = str(e)

            else:
                data['error'] = 'Accion no valida'

        except Exception as e:
            import traceback
            print("ERROR:", e)
            traceback.print_exc()
            data['error'] = str(e)

        return JsonResponse(data, safe=False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Copiar Precria desde Excel'
        # Piscinas de precria para el selector del formulario
        context['piscinas'] = Piscinas.objects.filter(prec__exact=True).order_by('orden')
        return context